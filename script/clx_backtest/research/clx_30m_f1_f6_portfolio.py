"""Build the locked CLX18 30-minute F1-F6 capital simulation and report.

The executable filter contract contains the 64 F1-F6 subsets, represented by
``filter_mask`` values in ``0..63``.

Inputs (all below ``--root``):

* ``features/candidate_events.parquet`` - immutable signal facts;
* ``matrix/locked_config.json`` - four TRAIN+VALIDATION locked selections;
* ``matrix/reveal_matrix.parquet`` - optional, already-once-revealed AUDIT facts;
* ``snapshot/bars/<code>.parquet`` - 30-minute QFQ closes used for marking;
* ``snapshot/index_day.parquet`` and ``features/market_segments.csv`` -
  market calendar/regime facts.

The simulator uses CNY 5,000,000, 40 slots with up to CNY 125,000 capital per
slot, exits before entries at an identical bar timestamp, never adds to an
already-held stock, and applies 0.02% on each side.  It intentionally does not
model slippage, stamp duty, minimum commission, or 100-share board-lot
rounding; those omissions are repeated in every generated report.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import tempfile
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import pyarrow.parquet as pq

if __package__:
    from . import clx_30m_f1_f6_matrix as _matrix_contract
else:
    import clx_30m_f1_f6_matrix as _matrix_contract  # type: ignore[no-redef]

_matrix_load_lock_for_reveal = _matrix_contract._load_lock_for_reveal
_matrix_reveal_identity = _matrix_contract._reveal_identity
_matrix_validate_reveal_lineage = _matrix_contract._validate_reveal_lineage

STUDY_ID = "clx-30m-full-trigger-f1-f6-v1"
PORTFOLIO_CONTRACT_VERSION = 1
INITIAL_CAPITAL = 5_000_000.0
MAX_POSITIONS = 40
SLOT_CAPITAL = 125_000.0
FEE_PER_SIDE = 0.0002
DAILY_ENTRY_LIMITS: tuple[int | None, ...] = (1, 3, 5, 10, 20, None)
HORIZONS = (5, 30, 60, 90)
CHECKPOINT_SCOPE_TOKENS = {
    "AVAILABLE": "available",
    "AUDIT": "audit",
    "MATCHED90": "matched90",
}
FILTER_NAMES = tuple(f"F{offset}" for offset in range(1, 7))
FILTER_DESCRIPTIONS = {
    "F1": "未复权原始开盘价1～6元",
    "F2": "个股近20个交易日收益≤0",
    "F3": "距近20个交易日高点回撤≥10%",
    "F4": "近20个交易日非年化日等效波动率≥3%",
    "F5": "收盘价≤MA60日等效均线",
    "F6": "冻结市场基准近20个完整交易日收益≤0",
}
TRIGGER_BITS = {
    "MODEL_STRUCTURAL": 0x01,
    "PIN_BAR": 0x02,
    "ENGULFING": 0x04,
    "STRONG_FRACTAL": 0x08,
    "MA5_TURN": 0x10,
    "PRICE_VOLUME_CONFIRMATION": 0x20,
    "MACD_CROSS": 0x40,
}
BAR_CLOCKS = (
    "10:00",
    "10:30",
    "11:00",
    "11:30",
    "13:30",
    "14:00",
    "14:30",
    "15:00",
)
SHANGHAI_TIMEZONE = "Asia/Shanghai"
DEFAULT_ROOT = Path(
    "D:/fqpack/runtime/clx-backtest/studies/clx-30m-full-trigger-f1-f6-v1"
)
REPRODUCE_SCRIPT = (
    r"<REPO_ROOT>\script\clx_backtest\research\clx_30m_f1_f6_portfolio.py"
)
PORTFOLIO_FRAME_NAMES = (
    "portfolio_summary",
    "random_order_runs",
    "random_order_sensitivity",
    "period_metrics",
    "equity_30m",
    "equity_daily",
    "chart_curves",
    "trades",
    "decision_summary",
    "locked_selections",
    "daily_baseline_comparison",
)
REQUIRED_LOGICAL_OUTPUTS = frozenset(
    {
        *(
            f"portfolio/{name}.{suffix}"
            for name in PORTFOLIO_FRAME_NAMES
            for suffix in ("parquet", "csv")
        ),
        "portfolio/clx_30m_portfolio_report.xlsx",
        "portfolio/report.md",
        "portfolio/portfolio_config.json",
        "portfolio/reproduce_command.txt",
    }
)

DAILY_BASELINES = {
    5: {
        "locked_combo": "MACD金叉+F4+F6",
        "train_net_win_rate": 0.5277,
        "validation_net_win_rate": 0.5719,
        "audit_net_win_rate": 0.4761,
        "total_return": -0.2324,
        "cagr": -0.0127,
        "max_drawdown": -0.4002,
    },
    30: {
        "locked_combo": "吞没+F3+F5",
        "train_net_win_rate": 0.5769,
        "validation_net_win_rate": 0.6135,
        "audit_net_win_rate": 0.5717,
        "total_return": 1.5508,
        "cagr": 0.0462,
        "max_drawdown": -0.4703,
    },
    60: {
        "locked_combo": "同时2个触发+F1+F2+F3",
        "train_net_win_rate": 0.6100,
        "validation_net_win_rate": 0.5995,
        "audit_net_win_rate": 0.6113,
        "total_return": 2.0773,
        "cagr": 0.0557,
        "max_drawdown": -0.6074,
    },
    90: {
        "locked_combo": "Pin Bar+F1+F4+F6",
        "train_net_win_rate": 0.6501,
        "validation_net_win_rate": 0.6141,
        "audit_net_win_rate": 0.7907,
        "total_return": 2.1358,
        "cagr": 0.0566,
        "max_drawdown": -0.5179,
    },
}


class PortfolioContractError(RuntimeError):
    """Raised when an immutable input or portfolio contract is invalid."""


@dataclass(frozen=True)
class LockedSelection:
    selection_id: str
    horizon: int
    model_codes: tuple[str, ...]
    trigger_kind: str
    trigger_value: int | None
    trigger_name: str
    trigger_id: str
    filter_mask: int
    filter_names: tuple[str, ...]
    development_score: float | None
    train_metrics: Mapping[str, Any]
    validation_metrics: Mapping[str, Any]

    def as_row(self) -> dict[str, Any]:
        return {
            "selection_id": self.selection_id,
            "horizon_trading_days": self.horizon,
            "model_code": ",".join(self.model_codes),
            "trigger_id": self.trigger_id,
            "trigger_kind": self.trigger_kind,
            "trigger_value": self.trigger_value,
            "trigger_name": self.trigger_name,
            "filter_mask": self.filter_mask,
            "filter_names": ",".join(self.filter_names),
            "development_score": self.development_score,
            "train_n": _metric(self.train_metrics, "sample_count", "n"),
            "train_net_win_rate": _metric(
                self.train_metrics, "net_win_rate", "win_rate"
            ),
            "train_mean_net_return": _metric(
                self.train_metrics, "mean_net_return", "net_mean_return"
            ),
            "validation_n": _metric(self.validation_metrics, "sample_count", "n"),
            "validation_net_win_rate": _metric(
                self.validation_metrics, "net_win_rate", "win_rate"
            ),
            "validation_mean_net_return": _metric(
                self.validation_metrics, "mean_net_return", "net_mean_return"
            ),
        }


@dataclass
class SimulationResult:
    summary: dict[str, Any]
    equity: pd.DataFrame
    trades: pd.DataFrame
    decisions: pd.DataFrame


def _metric(value: Mapping[str, Any], *names: str) -> Any:
    for name in names:
        if name in value:
            return value[name]
    return None


def _canonical_bytes(value: object) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        default=_json_default,
    ).encode("utf-8")


def _json_default(value: object) -> object:
    if isinstance(value, (pd.Timestamp, np.datetime64)):
        return pd.Timestamp(value).isoformat()
    if isinstance(value, np.generic):
        return value.item()
    if value is pd.NA:
        return None
    raise TypeError(f"{type(value).__name__} is not JSON serializable")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _atomic_json(path: Path, value: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(
        value,
        ensure_ascii=False,
        indent=2,
        sort_keys=True,
        default=_json_default,
    )
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as stream:
        stream.write(payload)
        temporary = Path(stream.name)
    os.replace(temporary, path)


def _atomic_parquet(frame: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as stream:
        temporary = Path(stream.name)
    try:
        frame.to_parquet(temporary, index=False)
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _atomic_csv(frame: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        mode="w",
        encoding="utf-8-sig",
        newline="",
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        delete=False,
    ) as stream:
        temporary = Path(stream.name)
    try:
        frame.to_csv(temporary, index=False, encoding="utf-8-sig")
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _as_shanghai(values: Iterable[object]) -> pd.DatetimeIndex:
    result = pd.DatetimeIndex(pd.to_datetime(values, errors="coerce", utc=True))
    return result.tz_convert(SHANGHAI_TIMEZONE)


def _normalise_model_codes(value: object) -> tuple[str, ...]:
    if isinstance(value, str):
        items = [item.strip() for item in value.split(",") if item.strip()]
    elif isinstance(value, Sequence):
        items = [str(item).strip() for item in value if str(item).strip()]
    else:
        items = []
    if not items:
        raise PortfolioContractError("locked selection has no model_code")
    for item in items:
        if item not in {"ALL", "*", "UNION"} and not (
            len(item) == 5 and item.startswith("S") and item[1:].isdigit()
        ):
            raise PortfolioContractError(f"invalid locked model_code: {item}")
    return tuple(items)


def _validated_matrix_lineage(
    root: Path,
) -> tuple[Path, dict[str, Any], Path, dict[str, Any]]:
    """Translate the Matrix fail-closed lineage contract for Portfolio callers."""

    try:
        lock_path, locked = _matrix_load_lock_for_reveal(root)
        config_path, study_config = _matrix_validate_reveal_lineage(
            root,
            lock_path=lock_path,
            locked=locked,
        )
    except (KeyError, OSError, TypeError, ValueError, RuntimeError) as exc:
        raise PortfolioContractError(f"matrix lock lineage mismatch: {exc}") from exc
    return lock_path, locked, config_path, study_config


def load_locked_selections(path: Path) -> list[LockedSelection]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except FileNotFoundError as exc:
        raise PortfolioContractError(f"locked config is missing: {path}") from exc
    if not isinstance(payload, Mapping) or payload.get("study_id") != STUDY_ID:
        raise PortfolioContractError("locked_config.json study_id mismatch")
    if path.name != "locked_config.json" or path.parent.name != "matrix":
        raise PortfolioContractError("locked_config.json path is outside matrix/")
    root = path.resolve().parents[1]
    validated_path, validated, _, _ = _validated_matrix_lineage(root)
    if validated_path.resolve() != path.resolve() or dict(payload) != validated:
        raise PortfolioContractError("locked_config.json lineage payload mismatch")
    return _parse_locked_selections(payload)


def _parse_locked_selections(payload: Mapping[str, Any]) -> list[LockedSelection]:
    filter_contract = payload.get("filter_contract")
    if isinstance(filter_contract, Mapping) and (
        int(filter_contract.get("subset_count", -1)) != 64
        or list(filter_contract.get("mask_range", [])) != [0, 63]
        or list(filter_contract.get("filters", [])) != list(FILTER_NAMES)
    ):
        raise PortfolioContractError(
            "locked_config.json F1-F6 filter contract mismatch"
        )
    raw = payload.get("selections") if isinstance(payload, Mapping) else None
    if not isinstance(raw, list):
        raise PortfolioContractError("locked_config.json must contain selections[]")
    selections: list[LockedSelection] = []
    for item in raw:
        if not isinstance(item, Mapping):
            raise PortfolioContractError("locked selection must be an object")
        selector = item.get("trigger_selector")
        if not isinstance(selector, Mapping):
            raise PortfolioContractError("locked trigger_selector is missing")
        kind = str(selector.get("kind", "")).upper()
        if kind not in {"SINGLE_BIT", "EXACT_MASK", "COUNT_EQ", "COUNT_GTE", "ALL"}:
            raise PortfolioContractError(f"unsupported trigger kind: {kind}")
        value = selector.get("value")
        parsed_value = None if value is None else int(value)
        if kind == "SINGLE_BIT" and parsed_value not in TRIGGER_BITS.values():
            raise PortfolioContractError(
                "SINGLE_BIT value must be a native trigger bit"
            )
        if kind == "EXACT_MASK" and not (1 <= int(parsed_value or 0) <= 127):
            raise PortfolioContractError("EXACT_MASK value must be 1..127")
        if kind == "COUNT_EQ" and parsed_value != 2:
            raise PortfolioContractError("COUNT_EQ is frozen to value=2")
        if kind == "COUNT_GTE" and parsed_value != 3:
            raise PortfolioContractError("COUNT_GTE is frozen to value=3")
        if kind == "ALL" and parsed_value is not None:
            raise PortfolioContractError("ALL trigger value must be null")
        horizon = int(item.get("horizon_trading_days", item.get("horizon", 0)))
        if horizon not in HORIZONS:
            raise PortfolioContractError(f"invalid horizon: {horizon}")
        filter_mask = int(item.get("filter_mask", -1))
        if not 0 <= filter_mask <= 63:
            raise PortfolioContractError(
                f"F1-F6 filter_mask must be in 0..63, got {filter_mask}"
            )
        expected_names = tuple(
            name for bit, name in enumerate(FILTER_NAMES) if filter_mask & (1 << bit)
        )
        supplied_names = tuple(str(value) for value in item.get("filter_names", []))
        if supplied_names and supplied_names != expected_names:
            raise PortfolioContractError(
                f"selection {item.get('selection_id')} filter_names disagree with mask"
            )
        selection_id = str(item.get("selection_id", "")).strip()
        if not selection_id:
            raise PortfolioContractError("locked selection_id is missing")
        score = item.get("development_score")
        selections.append(
            LockedSelection(
                selection_id=selection_id,
                horizon=horizon,
                model_codes=_normalise_model_codes(
                    item.get("model_code", item.get("model_codes"))
                ),
                trigger_kind=kind,
                trigger_value=parsed_value,
                trigger_name=str(selector.get("name", kind)),
                trigger_id=str(item.get("trigger_id", kind)),
                filter_mask=filter_mask,
                filter_names=expected_names,
                development_score=(
                    float(score) if score is not None and pd.notna(score) else None
                ),
                train_metrics=dict(item.get("train_metrics") or {}),
                validation_metrics=dict(item.get("validation_metrics") or {}),
            )
        )
    horizons = [selection.horizon for selection in selections]
    if len(selections) != 4 or sorted(horizons) != list(HORIZONS):
        raise PortfolioContractError(
            "F1-F6 lock must contain exactly one selection for each 5/30/60/90 horizon"
        )
    if len({item.selection_id for item in selections}) != len(selections):
        raise PortfolioContractError("locked selection_id values are not unique")
    return sorted(selections, key=lambda item: item.horizon)


def validate_reveal_inputs(root: Path, lock_path: Path) -> dict[str, tuple[Path, str]]:
    paths = {
        "matrix": root / "matrix" / "reveal_matrix.parquet",
        "summary": root / "matrix" / "reveal_summary.csv",
        "locked_detailed": root / "matrix" / "reveal_locked_detailed.parquet",
        "group_detail": root / "matrix" / "reveal_locked_group_detail.parquet",
    }
    manifest_path = root / "matrix" / "reveal_manifest.json"
    if not manifest_path.is_file() or any(
        not path.is_file() for path in paths.values()
    ):
        raise PortfolioContractError(
            "portfolio requires the completed matrix reveal summary, locked "
            "detail, group detail, and reveal manifest"
        )
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        raise PortfolioContractError("matrix reveal contract is unreadable") from exc
    validated_lock_path, lock, config_path, _ = _validated_matrix_lineage(root)
    if (
        validated_lock_path.resolve() != lock_path.resolve()
        or manifest.get("study_id") != STUDY_ID
        or manifest.get("stage") != "reveal"
        or manifest.get("lock_id") != lock.get("lock_id")
    ):
        raise PortfolioContractError("matrix reveal identity mismatch")
    identity = manifest.get("identity")
    if not isinstance(identity, Mapping):
        raise PortfolioContractError("matrix reveal identity payload is missing")
    try:
        minimum_reveal_samples = int(identity["minimum_reveal_samples"])
    except (KeyError, TypeError, ValueError) as exc:
        raise PortfolioContractError(
            "matrix reveal minimum_reveal_samples is invalid"
        ) from exc
    current_stage_id, current_identity = _matrix_reveal_identity(
        lock_path=lock_path,
        features_path=root / "features" / "candidate_events.parquet",
        config_path=config_path,
        index_path=root / "snapshot" / "index_day.parquet",
        snapshot_manifest_path=root / "snapshot" / "manifest.json",
        min_reveal_samples=minimum_reveal_samples,
    )
    if (
        dict(identity) != current_identity
        or manifest.get("stage_id") != current_stage_id
    ):
        raise PortfolioContractError("matrix reveal stage/input identity mismatch")
    outputs = manifest.get("outputs")
    if not isinstance(outputs, Mapping) or set(outputs) != set(paths):
        raise PortfolioContractError("matrix reveal frozen output set mismatch")
    validated: dict[str, tuple[Path, str]] = {}
    for key, path in paths.items():
        expected = outputs.get(key)
        if not isinstance(expected, Mapping):
            raise PortfolioContractError(
                f"matrix reveal {key} metadata must be an object"
            )
        digest = sha256_file(path)
        if (
            expected.get("file_sha256") != digest
            or Path(str(expected.get("path", ""))).resolve() != path.resolve()
            or expected.get("file_size") != path.stat().st_size
        ):
            raise PortfolioContractError(f"matrix reveal {key} identity mismatch")
        validated[key] = (path, digest)
    validated["manifest"] = (manifest_path, sha256_file(manifest_path))
    return validated


def _trigger_mask(frame: pd.DataFrame, selection: LockedSelection) -> pd.Series:
    masks = (
        pd.to_numeric(frame["concurrent_trigger_mask"], errors="coerce")
        .fillna(0)
        .astype("int64")
    )
    counts = (
        pd.to_numeric(frame["concurrent_trigger_count"], errors="coerce")
        .fillna(masks.map(int.bit_count))
        .astype("int64")
    )
    if selection.trigger_kind == "SINGLE_BIT":
        return masks.map(
            lambda value: bool(int(value) & int(selection.trigger_value or 0))
        )
    if selection.trigger_kind == "EXACT_MASK":
        return masks.eq(int(selection.trigger_value or 0))
    if selection.trigger_kind == "COUNT_EQ":
        return counts.eq(int(selection.trigger_value or 0))
    if selection.trigger_kind == "COUNT_GTE":
        return counts.ge(int(selection.trigger_value or 0))
    return pd.Series(True, index=frame.index)


def _event_ids(frame: pd.DataFrame) -> pd.Series:
    if "union_signal_id" in frame:
        supplied = frame["union_signal_id"].astype("string")
    elif "signal_fact_id" in frame:
        supplied = frame["signal_fact_id"].astype("string")
    else:
        supplied = pd.Series(pd.NA, index=frame.index, dtype="string")
    generated = [
        "sha256:"
        + hashlib.sha256(
            _canonical_bytes(
                {
                    "code": str(code),
                    "model_code": str(model),
                    "reveal_at": pd.Timestamp(reveal).isoformat(),
                }
            )
        ).hexdigest()
        for code, model, reveal in zip(
            frame["code"], frame["model_code"], frame["reveal_at"], strict=True
        )
    ]
    return supplied.fillna(pd.Series(generated, index=frame.index)).astype(str)


def select_locked_candidates(
    events: pd.DataFrame,
    selection: LockedSelection,
    *,
    scope: str = "AVAILABLE",
) -> pd.DataFrame:
    required = {
        "code",
        "model_code",
        "split_id",
        "reveal_at",
        "entry_at",
        "qfq_entry_open",
        "concurrent_trigger_mask",
        "filter_pass_mask",
        f"h{selection.horizon}_status",
        f"h{selection.horizon}_exit_at",
        f"h{selection.horizon}_gross_return",
        f"h{selection.horizon}_split_boundary_status",
    }
    missing = sorted(required - set(events.columns))
    if missing:
        raise PortfolioContractError(f"candidate events miss columns: {missing}")
    wildcard = bool(set(selection.model_codes) & {"ALL", "*", "UNION"})
    frame = (
        events.copy()
        if wildcard
        else events.loc[events["model_code"].isin(selection.model_codes)].copy()
    )
    for column in ("reveal_at", "entry_at", f"h{selection.horizon}_exit_at"):
        frame[column] = _as_shanghai(frame[column])
    pass_masks = (
        pd.to_numeric(frame["filter_pass_mask"], errors="coerce")
        .fillna(0)
        .astype("int64")
    )
    filters_pass = (pass_masks & selection.filter_mask).eq(selection.filter_mask)
    executable = (
        frame.get("entry_executable", pd.Series(True, index=frame.index))
        .fillna(False)
        .astype(bool)
        & frame.get("entry_status", pd.Series("OK", index=frame.index))
        .astype("string")
        .eq("OK")
        .fillna(False)
        & frame[f"h{selection.horizon}_status"].astype("string").eq("OK").fillna(False)
        & frame[f"h{selection.horizon}_split_boundary_status"]
        .astype("string")
        .eq("AVAILABLE")
        .fillna(False)
    )
    scope_name = scope.upper()
    split_ids = frame["split_id"].astype("string")
    mask = (
        filters_pass
        & executable
        & _trigger_mask(frame, selection)
        & split_ids.isin(("TRAIN", "VALIDATION", "AUDIT")).fillna(False)
    )
    if scope_name == "AUDIT":
        mask &= split_ids.eq("AUDIT").fillna(False)
    elif scope_name != "AVAILABLE":
        raise PortfolioContractError(f"unsupported portfolio scope: {scope}")
    frame = frame.loc[mask].copy()
    if frame.empty:
        return frame
    frame["candidate_id"] = _event_ids(frame)
    frame["exit_at"] = frame[f"h{selection.horizon}_exit_at"]
    frame["qfq_entry_open"] = pd.to_numeric(frame["qfq_entry_open"], errors="coerce")
    frame["gross_return"] = pd.to_numeric(
        frame[f"h{selection.horizon}_gross_return"], errors="coerce"
    )
    frame["qfq_exit_open"] = frame["qfq_entry_open"] * (1 + frame["gross_return"])
    frame = frame[
        frame["qfq_entry_open"].gt(0)
        & frame["qfq_exit_open"].gt(0)
        & frame["exit_at"].gt(frame["entry_at"])
    ].copy()
    quality_columns = {
        "concurrent_trigger_count": 0,
        "same_code_reveal_model_count": 1,
        "same_reveal_event_count": 1,
        "amount_median_20d": 0.0,
        "raw_entry_gap": 0.0,
    }
    for column, default in quality_columns.items():
        if column not in frame:
            frame[column] = default
        frame[column] = pd.to_numeric(frame[column], errors="coerce").fillna(default)
    frame = frame.sort_values(
        [
            "entry_at",
            "concurrent_trigger_count",
            "same_code_reveal_model_count",
            "same_reveal_event_count",
            "amount_median_20d",
            "raw_entry_gap",
            "candidate_id",
        ],
        ascending=[True, False, False, True, False, True, True],
        kind="stable",
    )
    # Union duplicate facts represent the same stock/reveal decision.
    return frame.drop_duplicates("candidate_id", keep="first").reset_index(drop=True)


class MarkStore:
    """Lazy, audited QFQ 30-minute close lookup."""

    def __init__(self, root: Path):
        self.root = root
        self._values: dict[str, tuple[np.ndarray, np.ndarray]] = {}
        self._aligned: dict[tuple[str, int], np.ndarray] = {}
        self.source_rows = 0

    def load(self, codes: Iterable[str]) -> None:
        for code in sorted({str(value) for value in codes}):
            if code in self._values:
                continue
            path = self.root / "snapshot" / "bars" / f"{code}.parquet"
            if not path.is_file():
                raise PortfolioContractError(
                    f"mark bars are missing for {code}: {path}"
                )
            frame = pd.read_parquet(path, columns=["bar_at", "qfq_close"])
            frame["bar_at"] = _as_shanghai(frame["bar_at"])
            frame["qfq_close"] = pd.to_numeric(frame["qfq_close"], errors="coerce")
            frame = (
                frame.dropna(subset=["bar_at", "qfq_close"])
                .loc[lambda value: value["qfq_close"].gt(0)]
                .sort_values("bar_at", kind="stable")
                .drop_duplicates("bar_at", keep="last")
            )
            if frame.empty:
                raise PortfolioContractError(
                    f"mark bars have no usable rows for {code}"
                )
            self._values[code] = (
                pd.DatetimeIndex(frame["bar_at"]).asi8,
                frame["qfq_close"].to_numpy(dtype=float),
            )
            self.source_rows += len(frame)

    def close(self, code: str, timestamp: pd.Timestamp, fallback: float) -> float:
        times, closes = self._values[code]
        offset = int(np.searchsorted(times, timestamp.value, side="right") - 1)
        return float(closes[offset]) if offset >= 0 else float(fallback)

    def aligned_closes(self, code: str, clock: pd.DatetimeIndex) -> np.ndarray:
        """Carry the last observed QFQ close across one immutable clock."""

        key = (code, id(clock))
        cached = self._aligned.get(key)
        if cached is not None:
            return cached
        times, closes = self._values[code]
        offsets = np.searchsorted(times, clock.asi8, side="right") - 1
        output = np.full(len(clock), np.nan, dtype=float)
        valid = offsets >= 0
        output[valid] = closes[offsets[valid]]
        self._aligned[key] = output
        return output

    def all_timestamps(self, codes: Iterable[str]) -> pd.DatetimeIndex:
        arrays = [self._values[str(code)][0] for code in sorted(set(codes))]
        if not arrays:
            return pd.DatetimeIndex([], tz=SHANGHAI_TIMEZONE)
        values = np.unique(np.concatenate(arrays))
        return pd.DatetimeIndex(values, tz="UTC").tz_convert(SHANGHAI_TIMEZONE)


def build_simulation_clock(
    root: Path,
    selected_frames: Sequence[pd.DataFrame],
    marks: MarkStore,
) -> pd.DatetimeIndex:
    nonempty = [frame for frame in selected_frames if not frame.empty]
    if not nonempty:
        raise PortfolioContractError(
            "all locked selections have zero portfolio candidates"
        )
    first = min(frame["entry_at"].min() for frame in nonempty)
    last = max(frame["exit_at"].max() for frame in nonempty)
    index_path = root / "snapshot" / "index_day.parquet"
    if index_path.is_file():
        sessions = pd.to_datetime(
            pd.read_parquet(index_path, columns=["date"])["date"],
            errors="coerce",
        ).dropna()
        dates = sorted(
            {
                value.date()
                for value in sessions
                if first.date() <= value.date() <= last.date()
            }
        )
        regular = pd.DatetimeIndex(
            [
                pd.Timestamp(f"{session.isoformat()} {clock}", tz=SHANGHAI_TIMEZONE)
                for session in dates
                for clock in BAR_CLOCKS
            ]
        )
    else:
        codes = {
            str(code)
            for frame in nonempty
            for code in frame["code"].astype(str).unique()
        }
        regular = marks.all_timestamps(codes)
        regular = regular[(regular >= first) & (regular <= last)]
    actions = pd.DatetimeIndex(
        [
            value
            for frame in nonempty
            for column in ("entry_at", "exit_at")
            for value in frame[column]
        ]
    )
    clock = regular.union(actions).sort_values().unique()
    clock = clock[(clock >= first) & (clock <= last)]
    if clock.empty:
        raise PortfolioContractError("simulation clock is empty")
    return clock


def _stable_order(frame: pd.DataFrame, policy: str, seed: int | None) -> pd.DataFrame:
    if len(frame) <= 1:
        return frame
    if policy == "quality":
        return frame.sort_values(
            [
                "concurrent_trigger_count",
                "same_code_reveal_model_count",
                "same_reveal_event_count",
                "amount_median_20d",
                "raw_entry_gap",
                "candidate_id",
            ],
            ascending=[False, False, True, False, True, True],
            kind="stable",
        )
    if policy != "sha_random" or seed is None:
        raise PortfolioContractError(f"invalid ranking policy: {policy}")
    ordered = frame.copy()
    ordered["_order_hash"] = [
        hashlib.sha256(f"{seed:03d}|{value}".encode()).hexdigest()
        for value in ordered["candidate_id"]
    ]
    return ordered.sort_values(["_order_hash", "candidate_id"], kind="stable").drop(
        columns="_order_hash"
    )


def prepare_ordered_entries(
    candidates: pd.DataFrame,
    *,
    ranking_policy: str,
    random_seed: int | None = None,
) -> dict[int, list[dict[str, Any]]]:
    """Freeze one causal within-timestamp ordering for reuse across six caps."""

    columns = (
        "candidate_id",
        "code",
        "qfq_entry_open",
        "qfq_exit_open",
        "exit_at",
        "market_regime",
    )
    output: dict[int, list[dict[str, Any]]] = {}
    for timestamp, group in candidates.groupby("entry_at", sort=False):
        ordered = _stable_order(group, ranking_policy, random_seed)
        output[pd.Timestamp(timestamp).value] = ordered.loc[:, columns].to_dict(
            orient="records"
        )
    return output


def _max_consecutive_losses(returns: Iterable[float]) -> int:
    longest = current = 0
    for value in returns:
        if value < 0:
            current += 1
            longest = max(longest, current)
        else:
            current = 0
    return longest


def _underwater_stats(timestamps: pd.Series, drawdowns: pd.Series) -> tuple[int, float]:
    start: pd.Timestamp | None = None
    longest_bars = 0
    current_bars = 0
    longest_days = 0.0
    for timestamp, value in zip(timestamps, drawdowns, strict=True):
        if float(value) < -1e-14:
            if start is None:
                start = pd.Timestamp(timestamp)
                current_bars = 0
            current_bars += 1
            longest_bars = max(longest_bars, current_bars)
            longest_days = max(
                longest_days,
                (pd.Timestamp(timestamp) - start).total_seconds() / 86_400,
            )
        else:
            start = None
            current_bars = 0
    return longest_bars, longest_days


def simulate_portfolio(
    candidates: pd.DataFrame,
    *,
    selection: LockedSelection,
    scope: str,
    clock: pd.DatetimeIndex,
    marks: MarkStore,
    daily_entry_limit: int | None,
    ranking_policy: str,
    random_seed: int | None = None,
    ordered_entries: Mapping[int, Sequence[Mapping[str, Any]]] | None = None,
    record_decisions: bool = False,
) -> SimulationResult:
    entries = (
        ordered_entries
        if ordered_entries is not None
        else prepare_ordered_entries(
            candidates,
            ranking_policy=ranking_policy,
            random_seed=random_seed,
        )
    )
    cash = INITIAL_CAPITAL
    positions: dict[str, dict[str, Any]] = {}
    opened_by_day: dict[object, int] = {}
    total_fees = 0.0
    buy_notional = 0.0
    sell_notional = 0.0
    peak_positions = 0
    decisions: list[dict[str, Any]] = []
    trades: list[dict[str, Any]] = []
    cash_values = np.empty(len(clock), dtype=float)
    position_counts = np.empty(len(clock), dtype=np.int16)
    rejected = {
        "occupied": 0,
        "daily_limit": 0,
        "slots": 0,
        "cash": 0,
    }

    def reject(row: Mapping[str, Any], timestamp: pd.Timestamp, reason: str) -> None:
        rejected[reason] += 1
        if record_decisions:
            decisions.append(
                {
                    "selection_id": selection.selection_id,
                    "horizon_trading_days": selection.horizon,
                    "scope": scope,
                    "daily_entry_limit": _limit_label(daily_entry_limit),
                    "ranking_policy": ranking_policy,
                    "random_seed": random_seed,
                    "candidate_id": row["candidate_id"],
                    "code": row["code"],
                    "entry_at": timestamp,
                    "decision": f"REJECT_{reason.upper()}",
                }
            )

    for clock_index, timestamp in enumerate(clock):
        timestamp = pd.Timestamp(timestamp)
        # Contract: every due exit is booked before any entry at the same bar.
        due = sorted(
            [
                (code, position)
                for code, position in positions.items()
                if position["exit_at"] <= timestamp
            ],
            key=lambda value: (value[1]["exit_at"], value[0]),
        )
        for code, position in due:
            exit_notional = position["units"] * position["exit_price"]
            exit_fee = exit_notional * FEE_PER_SIDE
            cash += exit_notional - exit_fee
            total_fees += exit_fee
            sell_notional += exit_notional
            net_pnl = exit_notional - exit_fee - position["entry_cost"]
            net_return = (exit_notional - exit_fee) / position["entry_cost"] - 1
            trades.append(
                {
                    "selection_id": selection.selection_id,
                    "horizon_trading_days": selection.horizon,
                    "scope": scope,
                    "daily_entry_limit": _limit_label(daily_entry_limit),
                    "ranking_policy": ranking_policy,
                    "random_seed": random_seed,
                    "candidate_id": position["candidate_id"],
                    "code": code,
                    "entry_at": position["entry_at"],
                    "exit_at": timestamp,
                    "entry_price": position["entry_price"],
                    "exit_price": position["exit_price"],
                    "units": position["units"],
                    "entry_notional": position["entry_notional"],
                    "exit_notional": exit_notional,
                    "entry_fee": position["entry_fee"],
                    "exit_fee": exit_fee,
                    "total_fee": position["entry_fee"] + exit_fee,
                    "net_pnl": net_pnl,
                    "net_return": net_return,
                    "entry_market_regime": position["market_regime"],
                    "_entry_clock_index": position["entry_clock_index"],
                    "_exit_clock_index": clock_index,
                }
            )
            positions.pop(code)

        incoming = entries.get(timestamp.value)
        if incoming is not None:
            day = timestamp.date()
            for row in incoming:
                code = str(row["code"])
                if code in positions:
                    reject(row, timestamp, "occupied")
                    continue
                if (
                    daily_entry_limit is not None
                    and opened_by_day.get(day, 0) >= daily_entry_limit
                ):
                    reject(row, timestamp, "daily_limit")
                    continue
                if len(positions) >= MAX_POSITIONS:
                    reject(row, timestamp, "slots")
                    continue
                capital_budget = min(SLOT_CAPITAL, cash)
                if capital_budget <= 0:
                    reject(row, timestamp, "cash")
                    continue
                entry_notional = capital_budget / (1 + FEE_PER_SIDE)
                entry_fee = capital_budget - entry_notional
                units = entry_notional / float(row["qfq_entry_open"])
                cash -= capital_budget
                total_fees += entry_fee
                buy_notional += entry_notional
                positions[code] = {
                    "candidate_id": row["candidate_id"],
                    "entry_at": timestamp,
                    "exit_at": pd.Timestamp(row["exit_at"]),
                    "entry_price": float(row["qfq_entry_open"]),
                    "exit_price": float(row["qfq_exit_open"]),
                    "units": units,
                    "entry_notional": entry_notional,
                    "entry_fee": entry_fee,
                    "entry_cost": capital_budget,
                    "market_regime": str(row.get("market_regime", "UNKNOWN")),
                    "entry_clock_index": clock_index,
                }
                opened_by_day[day] = opened_by_day.get(day, 0) + 1
                if record_decisions:
                    decisions.append(
                        {
                            "selection_id": selection.selection_id,
                            "horizon_trading_days": selection.horizon,
                            "scope": scope,
                            "daily_entry_limit": _limit_label(daily_entry_limit),
                            "ranking_policy": ranking_policy,
                            "random_seed": random_seed,
                            "candidate_id": row["candidate_id"],
                            "code": code,
                            "entry_at": timestamp,
                            "decision": "ACCEPT",
                        }
                    )
        if cash < -0.01 or len(positions) > MAX_POSITIONS:
            raise PortfolioContractError(
                f"portfolio invariant failed at {timestamp}: "
                f"cash={cash}, positions={len(positions)}"
            )
        peak_positions = max(peak_positions, len(positions))
        cash_values[clock_index] = cash
        position_counts[clock_index] = len(positions)
    if positions:
        raise PortfolioContractError(
            f"{len(positions)} positions remain after the last mature exit"
        )
    invested_values = np.zeros(len(clock), dtype=float)
    for trade in trades:
        start = int(trade["_entry_clock_index"])
        end = int(trade["_exit_clock_index"])
        prices = marks.aligned_closes(str(trade["code"]), clock)[start:end]
        if np.isnan(prices).any():
            prices = np.where(np.isnan(prices), float(trade["entry_price"]), prices)
        invested_values[start:end] += float(trade["units"]) * prices
    equity_values = cash_values + invested_values
    if (equity_values <= 0).any():
        bad = int(np.flatnonzero(equity_values <= 0)[0])
        raise PortfolioContractError(
            f"portfolio equity invariant failed at {clock[bad]}: "
            f"equity={equity_values[bad]}"
        )
    equity_frame = pd.DataFrame(
        {
            "selection_id": selection.selection_id,
            "horizon_trading_days": selection.horizon,
            "scope": scope,
            "daily_entry_limit": _limit_label(daily_entry_limit),
            "ranking_policy": ranking_policy,
            "random_seed": random_seed,
            "bar_at": clock,
            "cash": cash_values,
            "invested_value": invested_values,
            "equity": equity_values,
            "positions": position_counts,
            "capital_utilization": invested_values / equity_values,
        }
    )
    equity_frame["normalized_equity"] = equity_frame["equity"] / INITIAL_CAPITAL
    equity_frame["drawdown"] = (
        equity_frame["equity"] / equity_frame["equity"].cummax() - 1
    )
    trade_frame = pd.DataFrame(trades)
    if not trade_frame.empty:
        trade_frame = trade_frame.drop(
            columns=["_entry_clock_index", "_exit_clock_index"]
        )
    decision_frame = pd.DataFrame(
        decisions,
        columns=[
            "selection_id",
            "horizon_trading_days",
            "scope",
            "daily_entry_limit",
            "ranking_policy",
            "random_seed",
            "candidate_id",
            "code",
            "entry_at",
            "decision",
        ],
    )
    trade_returns = (
        trade_frame["net_return"].to_numpy(dtype=float)
        if not trade_frame.empty
        else np.asarray([], dtype=float)
    )
    pnl = (
        trade_frame["net_pnl"].to_numpy(dtype=float)
        if not trade_frame.empty
        else np.asarray([], dtype=float)
    )
    winners = trade_returns[trade_returns > 0]
    losers = trade_returns[trade_returns < 0]
    longest_bars, longest_days = _underwater_stats(
        equity_frame["bar_at"], equity_frame["drawdown"]
    )
    elapsed_years = max(
        (
            equity_frame["bar_at"].iloc[-1] - equity_frame["bar_at"].iloc[0]
        ).total_seconds()
        / (365.25 * 86_400),
        1 / 365.25,
    )
    final_equity = float(equity_frame["equity"].iloc[-1])
    accepted = len(trade_frame)
    average_equity = float(equity_frame["equity"].mean())
    summary = {
        "selection_id": selection.selection_id,
        "horizon_trading_days": selection.horizon,
        "scope": scope,
        "daily_entry_limit": _limit_label(daily_entry_limit),
        "ranking_policy": ranking_policy,
        "random_seed": random_seed,
        "order_sha256": (
            "sha256:"
            + hashlib.sha256(
                f"{selection.selection_id}|{ranking_policy}|{random_seed}".encode()
            ).hexdigest()
        ),
        "initial_capital": INITIAL_CAPITAL,
        "final_equity": final_equity,
        "total_return": final_equity / INITIAL_CAPITAL - 1,
        "cagr": (final_equity / INITIAL_CAPITAL) ** (1 / elapsed_years) - 1,
        "max_drawdown": float(equity_frame["drawdown"].min()),
        "longest_underwater_bars": longest_bars,
        "longest_underwater_days": longest_days,
        "candidate_signals": len(candidates),
        "closed_trades": accepted,
        "candidate_acceptance_rate": (
            accepted / len(candidates) if len(candidates) else None
        ),
        "closed_win_rate": (
            float(np.mean(trade_returns > 0)) if len(trade_returns) else None
        ),
        "mean_trade_return": (
            float(trade_returns.mean()) if len(trade_returns) else None
        ),
        "median_trade_return": (
            float(np.median(trade_returns)) if len(trade_returns) else None
        ),
        "average_win": float(winners.mean()) if len(winners) else None,
        "average_loss_abs": float(abs(losers.mean())) if len(losers) else None,
        "payoff_ratio": (
            float(winners.mean() / abs(losers.mean()))
            if len(winners) and len(losers)
            else None
        ),
        "profit_factor": (
            float(pnl[pnl > 0].sum() / abs(pnl[pnl < 0].sum()))
            if np.any(pnl > 0) and np.any(pnl < 0)
            else None
        ),
        "max_consecutive_losses": _max_consecutive_losses(trade_returns),
        "total_fees": total_fees,
        "buy_notional": buy_notional,
        "sell_notional": sell_notional,
        "one_way_turnover": (
            (buy_notional + sell_notional) / (2 * average_equity)
            if average_equity > 0
            else None
        ),
        "two_way_turnover": (
            (buy_notional + sell_notional) / average_equity
            if average_equity > 0
            else None
        ),
        "average_positions": float(equity_frame["positions"].mean()),
        "peak_positions": peak_positions,
        "average_capital_utilization": float(
            equity_frame["capital_utilization"].mean()
        ),
        "peak_capital_utilization": float(equity_frame["capital_utilization"].max()),
        "rejected_occupied": rejected["occupied"],
        "rejected_daily_limit": rejected["daily_limit"],
        "rejected_slots": rejected["slots"],
        "rejected_cash": rejected["cash"],
        "start_at": equity_frame["bar_at"].iloc[0],
        "end_at": equity_frame["bar_at"].iloc[-1],
    }
    return SimulationResult(summary, equity_frame, trade_frame, decision_frame)


def _limit_label(value: int | None) -> str:
    return "UNLIMITED" if value is None else str(value)


def _regime_for_dates(dates: pd.Series, segments: pd.DataFrame) -> pd.Series:
    output = pd.Series("UNKNOWN", index=dates.index, dtype="string")
    if segments.empty:
        return output
    parsed = segments.copy()
    parsed["start_date"] = pd.to_datetime(parsed["start_date"]).dt.date
    parsed["end_date"] = pd.to_datetime(parsed["end_date"]).dt.date
    day_values = pd.to_datetime(dates).dt.date
    for row in parsed.itertuples(index=False):
        in_segment = day_values.ge(row.start_date) & day_values.le(row.end_date)
        output.loc[in_segment] = str(row.regime)
    return output


def build_period_metrics(
    equity: pd.DataFrame,
    trades: pd.DataFrame,
    segments: pd.DataFrame,
) -> pd.DataFrame:
    if equity.empty:
        return pd.DataFrame()
    frame = equity.sort_values("bar_at", kind="stable").copy()
    frame["year"] = frame["bar_at"].dt.year.astype(str)
    frame["quarter"] = (
        frame["bar_at"].dt.tz_localize(None).dt.to_period("Q").astype(str)
    )
    frame["regime"] = _regime_for_dates(frame["bar_at"], segments)
    frame["bar_return"] = (
        frame["equity"]
        .pct_change(fill_method=None)
        .fillna(frame["equity"].iloc[0] / INITIAL_CAPITAL - 1)
    )
    trade_frame = trades.copy()
    if not trade_frame.empty:
        trade_frame["year"] = trade_frame["entry_at"].dt.year.astype(str)
        trade_frame["quarter"] = (
            trade_frame["entry_at"].dt.tz_localize(None).dt.to_period("Q").astype(str)
        )
        trade_frame["regime"] = trade_frame["entry_market_regime"].fillna("UNKNOWN")
    keys = [
        "selection_id",
        "horizon_trading_days",
        "scope",
        "daily_entry_limit",
        "ranking_policy",
    ]
    identity = {key: frame[key].iloc[0] for key in keys}
    rows: list[dict[str, Any]] = []
    for period_type, column in (
        ("YEAR", "year"),
        ("QUARTER", "quarter"),
        ("REGIME", "regime"),
    ):
        for period_id, group in frame.groupby(column, sort=True):
            period_trades = (
                trade_frame[trade_frame[column].eq(period_id)]
                if not trade_frame.empty
                else trade_frame
            )
            returns = (
                period_trades["net_return"].to_numpy(dtype=float)
                if not period_trades.empty
                else np.asarray([], dtype=float)
            )
            rows.append(
                {
                    **identity,
                    "period_type": period_type,
                    "period_id": str(period_id),
                    "start_at": group["bar_at"].min(),
                    "end_at": group["bar_at"].max(),
                    "start_equity": float(group["equity"].iloc[0]),
                    "end_equity": float(group["equity"].iloc[-1]),
                    "normalized_end_equity": float(
                        group["equity"].iloc[-1] / INITIAL_CAPITAL
                    ),
                    "portfolio_return": float(
                        np.prod(1 + group["bar_return"].to_numpy(dtype=float)) - 1
                    ),
                    "max_drawdown_within_period": float(
                        (group["equity"] / group["equity"].cummax() - 1).min()
                    ),
                    "closed_trades": len(period_trades),
                    "closed_win_rate": (
                        float(np.mean(returns > 0)) if len(returns) else None
                    ),
                    "mean_trade_return": (
                        float(np.mean(returns)) if len(returns) else None
                    ),
                }
            )
    return pd.DataFrame(rows)


def summarise_random_runs(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    metrics = (
        "total_return",
        "cagr",
        "max_drawdown",
        "closed_win_rate",
        "profit_factor",
        "total_fees",
        "average_capital_utilization",
        "candidate_acceptance_rate",
    )
    keys = [
        "selection_id",
        "horizon_trading_days",
        "scope",
        "daily_entry_limit",
    ]
    rows: list[dict[str, Any]] = []
    for identity, group in frame.groupby(keys, dropna=False, sort=True):
        row = dict(zip(keys, identity, strict=True))
        row["random_runs"] = len(group)
        for metric in metrics:
            values = pd.to_numeric(group[metric], errors="coerce").dropna()
            for label, quantile in (("p05", 0.05), ("p50", 0.50), ("p95", 0.95)):
                row[f"{metric}_{label}"] = (
                    float(values.quantile(quantile)) if len(values) else None
                )
        rows.append(row)
    return pd.DataFrame(rows)


def build_daily_baseline_comparison(summary: pd.DataFrame) -> pd.DataFrame:
    selected = summary[
        summary["scope"].eq("AVAILABLE")
        & summary["daily_entry_limit"].eq("5")
        & summary["ranking_policy"].eq("quality")
    ].copy()
    rows: list[dict[str, Any]] = []
    for row in selected.itertuples(index=False):
        baseline = DAILY_BASELINES[int(row.horizon_trading_days)]
        rows.append(
            {
                "horizon_trading_days": row.horizon_trading_days,
                "selection_id_30m": row.selection_id,
                "daily_locked_combo": baseline["locked_combo"],
                "portfolio_total_return_30m": row.total_return,
                "portfolio_total_return_daily": baseline["total_return"],
                "total_return_delta_30m_minus_daily": (
                    row.total_return - baseline["total_return"]
                ),
                "portfolio_cagr_30m": row.cagr,
                "portfolio_cagr_daily": baseline["cagr"],
                "cagr_delta_30m_minus_daily": row.cagr - baseline["cagr"],
                "portfolio_mdd_30m": row.max_drawdown,
                "portfolio_mdd_daily": baseline["max_drawdown"],
                "mdd_delta_30m_minus_daily": row.max_drawdown
                - baseline["max_drawdown"],
                "daily_train_net_win_rate": baseline["train_net_win_rate"],
                "daily_validation_net_win_rate": baseline["validation_net_win_rate"],
                "daily_audit_net_win_rate": baseline["audit_net_win_rate"],
                "period_comparability": "DIFFERENT_SAMPLE_WINDOWS_REVIEW_REQUIRED",
            }
        )
    return pd.DataFrame(rows)


def _daily_curves(equity: pd.DataFrame) -> pd.DataFrame:
    if equity.empty:
        return equity
    frame = equity.copy()
    frame["trade_date"] = frame["bar_at"].dt.date
    keys = [
        "selection_id",
        "horizon_trading_days",
        "scope",
        "daily_entry_limit",
        "ranking_policy",
        "trade_date",
    ]
    return (
        frame.sort_values("bar_at", kind="stable")
        .groupby(keys, as_index=False, dropna=False)
        .tail(1)
        .reset_index(drop=True)
    )


def build_chart_data(daily_curves: pd.DataFrame) -> pd.DataFrame:
    selected = daily_curves[
        daily_curves["scope"].eq("AVAILABLE")
        & daily_curves["daily_entry_limit"].eq("5")
        & daily_curves["ranking_policy"].eq("quality")
    ].copy()
    if selected.empty:
        return pd.DataFrame()
    output: pd.DataFrame | None = None
    for metric in ("normalized_equity", "drawdown"):
        wide = selected.pivot_table(
            index="trade_date",
            columns="horizon_trading_days",
            values=metric,
            aggfunc="last",
        )
        wide = wide.rename(
            columns={horizon: f"h{int(horizon)}_{metric}" for horizon in wide.columns}
        ).reset_index()
        output = (
            wide
            if output is None
            else output.merge(wide, on="trade_date", how="outer", validate="one_to_one")
        )
    assert output is not None
    output["trade_date"] = pd.to_datetime(output["trade_date"])
    return output.sort_values("trade_date", kind="stable").reset_index(drop=True)


def _write_excel(
    path: Path,
    *,
    summary: pd.DataFrame,
    random_summary: pd.DataFrame,
    period_metrics: pd.DataFrame,
    daily_curves: pd.DataFrame,
    chart_data: pd.DataFrame,
    locked: pd.DataFrame,
    baseline: pd.DataFrame,
    reveal: pd.DataFrame,
    reveal_detailed: pd.DataFrame,
    reveal_groups: pd.DataFrame,
) -> None:
    def column_kind(column: str) -> str:
        name = column.lower()
        if name == "trade_date" or name.endswith("_date"):
            return "date"
        if name.endswith("_at"):
            return "datetime"
        if (
            "return" in name
            or "win_rate" in name
            or "cagr" in name
            or "drawdown" in name
            or "utilization" in name
            or "utilisation" in name
            or "acceptance_rate" in name
            or "normalized_" in name
            or name in {"average_win", "average_loss_abs"}
        ):
            return "percentage"
        if any(
            token in name
            for token in ("capital", "equity", "cash", "notional", "total_fee")
        ):
            return "currency"
        if (
            name in {"n", "random_seed", "random_runs", "filter_mask", "trigger_value"}
            or name.endswith(
                ("_n", "_count", "_rows", "_signals", "_trades", "_positions")
            )
            or name.startswith(("rejected_", "unique_"))
            or "horizon_trading_days" in name
        ):
            if name == "average_positions":
                return "decimal"
            return "integer"
        if (
            any(
                token in name
                for token in (
                    "profit_factor",
                    "payoff_ratio",
                    "turnover",
                    "average_loss",
                )
            )
            or name == "longest_underwater_days"
        ):
            return "decimal"
        if name.endswith("_id") or "sha256" in name:
            return "identifier"
        if any(
            token in name
            for token in (
                "name",
                "contract",
                "description",
                "warning",
                "reason",
                "comparability",
                "model_code",
                "filter_names",
            )
        ):
            return "text"
        return "general"

    def style_worksheet(worksheet: Any) -> None:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        if worksheet.max_column < 1:
            return
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        worksheet.row_dimensions[1].height = 34
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        for column_number, header in enumerate(worksheet[1], start=1):
            name = str(header.value or "")
            kind = column_kind(name)
            if kind == "identifier":
                width = 24
            elif kind == "text":
                sample_lengths = [
                    len(str(worksheet.cell(row, column_number).value or ""))
                    for row in range(2, min(worksheet.max_row, 101) + 1)
                ]
                width = min(
                    42,
                    max(16, len(name) + 2, max(sample_lengths, default=0) + 2),
                )
            elif kind == "datetime":
                width = 20
            elif kind == "date":
                width = 13
            elif kind == "currency":
                width = 18
            elif kind in {"percentage", "decimal"}:
                width = 15
            elif kind == "integer":
                width = 14
            else:
                width = min(24, max(12, len(name) + 2))
            worksheet.column_dimensions[get_column_letter(column_number)].width = width
            number_format = {
                "date": "yyyy-mm-dd",
                "datetime": "yyyy-mm-dd hh:mm",
                "percentage": "0.00%",
                "currency": '"¥"#,##0.00',
                "integer": "#,##0",
                "decimal": "0.00",
            }.get(kind)
            if number_format is not None:
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(row, column_number).number_format = number_format

    def excel_safe(frame: pd.DataFrame) -> pd.DataFrame:
        value = frame.copy()
        for column in value.columns:
            if isinstance(value[column].dtype, pd.DatetimeTZDtype):
                value[column] = value[column].dt.tz_localize(None)
        return value

    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.stem}.tmp{path.suffix}")
    temporary.unlink(missing_ok=True)
    excel_chart_data = excel_safe(chart_data)
    if "trade_date" in excel_chart_data:
        excel_chart_data["trade_date_label"] = pd.to_datetime(
            excel_chart_data["trade_date"],
            errors="coerce",
        ).dt.strftime("%Y-%m-%d")
    try:
        with pd.ExcelWriter(temporary, engine="openpyxl") as writer:
            excel_safe(summary).to_excel(writer, sheet_name="Portfolio", index=False)
            excel_safe(random_summary).to_excel(
                writer, sheet_name="RandomSensitivity", index=False
            )
            excel_safe(period_metrics).to_excel(
                writer, sheet_name="AnnualQuarterRegime", index=False
            )
            excel_safe(daily_curves).to_excel(
                writer, sheet_name="NormalizedCurves", index=False
            )
            excel_chart_data.to_excel(writer, sheet_name="ChartData", index=False)
            excel_safe(locked).to_excel(writer, sheet_name="LockedConfigs", index=False)
            excel_safe(baseline).to_excel(
                writer, sheet_name="DailyBaseline", index=False
            )
            if not reveal.empty:
                excel_safe(reveal.iloc[:1_048_575]).to_excel(
                    writer, sheet_name="AuditReveal", index=False
                )
            excel_safe(reveal_detailed.iloc[:1_048_575]).to_excel(
                writer, sheet_name="RevealDetailed", index=False
            )
            excel_safe(reveal_groups.iloc[:1_048_575]).to_excel(
                writer, sheet_name="RevealByPeriod", index=False
            )
            disclosure = pd.DataFrame(
                {
                    "item": [
                        "signal",
                        "entry",
                        "exit",
                        "fee",
                        "mark",
                        "not_modelled",
                    ],
                    "contract": [
                        "30分钟K线收盘揭示；只用当时及以前数据",
                        "下一根实际存在的30分钟K线开盘",
                        "第5/30/60/90个股票交易日同槽位，缺K线按候选事实的下一可交易K线",
                        "买入、卖出各0.02%",
                        "30分钟前复权收盘，缺当前K线时沿用此前最近收盘",
                        "滑点、印花税、最低佣金、100股取整",
                    ],
                }
            )
            disclosure.to_excel(writer, sheet_name="Contract", index=False)
            charts = writer.book.create_sheet("Charts")
            if not chart_data.empty:
                from openpyxl.chart import LineChart, Reference
                from openpyxl.chart.data_source import AxDataSource, StrRef
                from openpyxl.chart.series import SeriesLabel

                data_sheet = writer.book["ChartData"]
                date_label_column = (
                    excel_chart_data.columns.get_loc("trade_date_label") + 1
                )
                equity_columns = [
                    offset
                    for offset, name in enumerate(chart_data.columns, start=1)
                    if str(name).endswith("_normalized_equity")
                ]
                drawdown_columns = [
                    offset
                    for offset, name in enumerate(chart_data.columns, start=1)
                    if str(name).endswith("_drawdown")
                ]
                for title, columns, anchor, y_title in (
                    (
                        "5/30/60/90日归一净值（每日上限5）",
                        equity_columns,
                        "A1",
                        "归一净值",
                    ),
                    (
                        "5/30/60/90日回撤（每日上限5）",
                        drawdown_columns,
                        "A20",
                        "回撤",
                    ),
                ):
                    chart = LineChart()
                    chart.title = title
                    chart.y_axis.title = y_title
                    chart.y_axis.numFmt = "0.00%"
                    chart.x_axis.title = "交易日"
                    chart.height = 8
                    chart.width = 18
                    for column in columns:
                        column_name = str(chart_data.columns[column - 1])
                        horizon = column_name.split("_", 1)[0].removeprefix("h")
                        chart.add_data(
                            Reference(
                                data_sheet,
                                min_col=column,
                                max_col=column,
                                min_row=1,
                                max_row=len(chart_data) + 1,
                            ),
                            titles_from_data=True,
                        )
                        chart.series[-1].tx = SeriesLabel(v=f"{horizon}日")
                    categories = Reference(
                        data_sheet,
                        min_col=date_label_column,
                        min_row=2,
                        max_row=len(chart_data) + 1,
                    )
                    for series in chart.series:
                        series.cat = AxDataSource(strRef=StrRef(f=str(categories)))
                    charts.add_chart(chart, anchor)
            for worksheet in writer.book.worksheets:
                style_worksheet(worksheet)
                if worksheet.max_row > 1:
                    worksheet.freeze_panes = "A2"
                    worksheet.auto_filter.ref = worksheet.dimensions
            from openpyxl.styles import Alignment

            contract_sheet = writer.book["Contract"]
            for row in range(2, contract_sheet.max_row + 1):
                contract_sheet.cell(row, 2).alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
                contract_sheet.row_dimensions[row].height = 36
            charts.sheet_view.showGridLines = False
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _fmt_pct(value: Any) -> str:
    if value is None or pd.isna(value):
        return "NA"
    return f"{float(value):.2%}"


def _fmt_ratio(value: Any) -> str:
    if value is None or pd.isna(value):
        return "NA"
    number = float(value)
    return "∞" if np.isposinf(number) else f"{number:.2f}"


def _portfolio_quality_statement(row: pd.Series) -> str:
    total_return = float(row["total_return"])
    raw_profit_factor = row["profit_factor"]
    profit_factor = (
        None
        if raw_profit_factor is None or pd.isna(raw_profit_factor)
        else float(raw_profit_factor)
    )
    if profit_factor is None:
        interpretation = "PF缺少可计算的已平仓盈亏样本"
    elif total_return > 0 and profit_factor > 1:
        interpretation = "总收益与PF均高于各自基准"
    elif total_return <= 0 and profit_factor <= 1:
        interpretation = "总收益与PF均未高于各自基准"
    else:
        interpretation = "总收益方向与PF基准不一致"
    return (
        f"总收益 `{_fmt_pct(total_return)}`、PF `{_fmt_ratio(profit_factor)}`、"
        f"平仓胜率 `{_fmt_pct(row['closed_win_rate'])}`；{interpretation}。"
    )


def _stability_statement(
    available: pd.Series,
    audit: pd.Series | None,
) -> str:
    if audit is None:
        return (
            f"AVAILABLE总收益 `{_fmt_pct(available['total_return'])}`、"
            f"PF `{_fmt_ratio(available['profit_factor'])}`；资金AUDIT未执行。"
        )
    available_return = float(available["total_return"])
    audit_return = float(audit["total_return"])
    raw_profit_factor = audit["profit_factor"]
    audit_profit_factor = (
        None
        if raw_profit_factor is None or pd.isna(raw_profit_factor)
        else float(raw_profit_factor)
    )
    if available_return > 0 and audit_return > 0:
        return_interpretation = "两段总收益均为正，资金方向一致"
    elif available_return < 0 and audit_return < 0:
        return_interpretation = "两段总收益均为负，资金方向一致"
    elif available_return * audit_return < 0:
        return_interpretation = "AVAILABLE与AUDIT总收益异号，资金方向不一致"
    elif available_return == 0 and audit_return == 0:
        return_interpretation = "两段总收益均为零"
    else:
        return_interpretation = "至少一段总收益为零，资金方向未完全一致"
    if audit_profit_factor is None:
        pf_interpretation = "AUDIT PF缺少可计算的已平仓盈亏样本"
    elif audit_profit_factor > 1:
        pf_interpretation = "AUDIT PF高于1"
    elif audit_profit_factor == 1:
        pf_interpretation = "AUDIT PF等于1"
    else:
        pf_interpretation = "AUDIT PF未高于1"
    return (
        f"AVAILABLE总收益 `{_fmt_pct(available_return)}`、"
        f"AUDIT总收益 `{_fmt_pct(audit_return)}`、"
        f"AUDIT PF `{_fmt_ratio(audit_profit_factor)}`、"
        f"AUDIT平仓胜率 `{_fmt_pct(audit['closed_win_rate'])}`；"
        f"{return_interpretation}；{pf_interpretation}。"
    )


def _audit_period_extrema(period_metrics: pd.DataFrame) -> pd.DataFrame:
    required = {
        "scope",
        "daily_entry_limit",
        "ranking_policy",
        "horizon_trading_days",
        "period_type",
        "period_id",
        "portfolio_return",
        "closed_trades",
        "closed_win_rate",
    }
    if period_metrics.empty or not required.issubset(period_metrics.columns):
        return pd.DataFrame()
    audit = period_metrics[
        period_metrics["scope"].eq("AUDIT")
        & period_metrics["daily_entry_limit"].eq("5")
        & period_metrics["ranking_policy"].eq("quality")
    ].copy()
    audit["portfolio_return"] = pd.to_numeric(
        audit["portfolio_return"], errors="coerce"
    )
    audit = audit.dropna(subset=["portfolio_return"])
    rows: list[dict[str, Any]] = []
    for identity, group in audit.groupby(
        ["horizon_trading_days", "period_type"],
        sort=True,
    ):
        horizon, period_type = identity
        ordered = group.sort_values(
            ["portfolio_return", "period_id"],
            ascending=[False, True],
            kind="stable",
        )
        best = ordered.iloc[0]
        worst = ordered.iloc[-1]
        rows.append(
            {
                "horizon_trading_days": int(horizon),
                "period_type": str(period_type),
                "best_period_id": str(best["period_id"]),
                "best_return": float(best["portfolio_return"]),
                "best_closed_trades": int(best["closed_trades"]),
                "best_closed_win_rate": best["closed_win_rate"],
                "worst_period_id": str(worst["period_id"]),
                "worst_return": float(worst["portfolio_return"]),
            }
        )
    return pd.DataFrame(rows)


def build_markdown_report(
    *,
    root: Path,
    summary: pd.DataFrame,
    random_summary: pd.DataFrame,
    period_metrics: pd.DataFrame,
    locked: pd.DataFrame,
    baseline: pd.DataFrame,
    feature_summary: Mapping[str, Any],
    reveal_summary: pd.DataFrame,
    reveal_detailed: pd.DataFrame,
    index_benchmark: Mapping[str, Any],
    portfolio_logic_sha256: str,
) -> str:
    dates = summary[["start_at", "end_at"]].copy()
    minimum = pd.to_datetime(dates["start_at"]).min()
    maximum = pd.to_datetime(dates["end_at"]).max()
    short_sample = minimum.year >= 2024
    grade = "SHORT_SAMPLE（短样本）" if short_sample else "FULL_HISTORY"
    random_run_counts = pd.to_numeric(
        random_summary.get("random_runs", pd.Series(dtype=float)),
        errors="coerce",
    ).dropna()
    unique_random_run_counts = sorted(
        {int(value) for value in random_run_counts if value >= 0}
    )
    if not unique_random_run_counts:
        random_contract_text = "本次未执行SHA确定性随机排序敏感性。"
        random_output_text = "本次未生成SHA随机排序分位数"
    else:
        random_run_label = (
            f"{unique_random_run_counts[0]}组"
            if len(unique_random_run_counts) == 1
            else (
                f"{unique_random_run_counts[0]}至" f"{unique_random_run_counts[-1]}组"
            )
        )
        random_contract_text = (
            f"{random_run_label}敏感性用 " "`SHA256(seed|candidate_id)` 确定排序。"
        )
        random_output_text = f"{random_run_label}SHA随机排序分位数"
    benchmark_label = str(index_benchmark["benchmark_label"])
    benchmark_role = "ETF代理" if bool(index_benchmark["is_proxy"]) else "指数基准"
    lines = [
        "# CLX18 30分钟锁定组合真实资金回测",
        "",
        "## 一、数据事实",
        "",
        f"- 证据等级：**{grade}**。",
        f"- 资金曲线区间：`{minimum.isoformat()}` 至 `{maximum.isoformat()}`。",
        (
            f"- 候选事件数：`{feature_summary.get('candidate_event_rows', 'NA')}`；"
            f"Union去重信号数：`{feature_summary.get('unique_union_signals', 'NA')}`。"
        ),
        f"- AUDIT揭示摘要：`{'已读取' if not reveal_summary.empty else '本次报告未找到揭示产物'}`。",
        "- 30分钟数据由本地 MongoDB 冻结为不可变 snapshot；报告阶段不修改源数据。",
        (
            f"- 市场基准：**{benchmark_label}**（{benchmark_role}；"
            f"`source_kind={index_benchmark['source_kind']}`，"
            f"`source_name={index_benchmark['source_name']}`）；"
            "报告内全部超额收益均相对此冻结基准计算。"
        ),
        "",
        "## 二、冻结研究假设与资金合同",
        "",
        "- 模型冠军只由 TRAIN+VALIDATION 锁定；资金模拟直接使用 `matrix/locked_config.json`，不按 AUDIT 或资金结果重选。",
        "- 过滤空间为 **F1-F6 共64个子集**，`filter_mask` 范围 `0..63`。",
        "- 初始资金500万元；40槽；每槽最多12.5万元（含买入费的资本预算）。",
        "- 同一时点先退出后入场；同股持有期间不加仓；每日新开上限分别为1/3/5/10/20/不限。",
        (
            "- 同一可交易时点内才做质量排序；不跨未来时点重排。"
            f"{random_contract_text}"
        ),
        "- 买卖各收0.02%；30分钟前复权收盘盯市，停牌时沿用此前最近可得前复权收盘。",
        "- 未计：**滑点、印花税、最低佣金、100股取整**。",
        "",
        "## 三、样本内锁定配置",
        "",
        "|期限|模型|触发|过滤|TRAIN n/胜率|VALIDATION n/胜率|",
        "|---:|---|---|---|---:|---:|",
    ]
    for row in locked.itertuples(index=False):
        lines.append(
            f"|{row.horizon_trading_days}|{row.model_code}|{row.trigger_name}|"
            f"{row.filter_names or '零过滤'}|"
            f"{row.train_n}/{_fmt_pct(row.train_net_win_rate)}|"
            f"{row.validation_n}/{_fmt_pct(row.validation_net_win_rate)}|"
        )
    lines.extend(
        [
            "",
            "## 四、样本外 AUDIT 一次性揭示",
            "",
            (
                "|期限|n|净胜率|95% CI|平均净收益|中位净收益|PF|"
                f"相对{benchmark_label}平均超额|提示|"
            ),
            "|---:|---:|---:|---:|---:|---:|---:|---:|---|",
        ]
    )
    audit = (
        reveal_summary[reveal_summary["scope"].eq("AUDIT")]
        .sort_values("horizon_trading_days")
        .copy()
        if not reveal_summary.empty and "scope" in reveal_summary
        else pd.DataFrame()
    )
    if audit.empty:
        lines.append("|—|—|—|—|—|—|—|—|揭示摘要尚未产出|")
    else:
        for row in audit.itertuples(index=False):
            warning = (
                "小样本；多重比较"
                if bool(getattr(row, "small_sample_warning", False))
                else "多重比较"
            )
            lines.append(
                f"|{row.horizon_trading_days}|{row.sample_count}|"
                f"{_fmt_pct(row.net_win_rate)}|"
                f"[{_fmt_pct(row.net_win_rate_ci_low)},"
                f"{_fmt_pct(row.net_win_rate_ci_high)}]|"
                f"{_fmt_pct(row.mean_net_return)}|"
                f"{_fmt_pct(row.median_net_return)}|"
                f"{'NA' if pd.isna(row.profit_factor) else f'{row.profit_factor:.2f}'}|"
                f"{_fmt_pct(row.mean_net_excess_return)}|{warning}|"
            )
    exact_audit = (
        reveal_detailed[reveal_detailed["scope"].eq("AUDIT")]
        .sort_values(
            ["horizon_trading_days", "model_population", "aggregation"],
            kind="stable",
        )
        .copy()
        if not reveal_detailed.empty
        and {
            "scope",
            "model_population",
            "aggregation",
        }.issubset(reveal_detailed.columns)
        else pd.DataFrame()
    )
    lines.extend(
        [
            "",
            "### AUDIT多聚合口径",
            "",
            "|期限|模型总体|聚合|n|净胜率|平均净收益|平均超额|",
            "|---:|---|---|---:|---:|---:|---:|",
        ]
    )
    if exact_audit.empty:
        lines.append("|—|—|—|—|—|—|揭示明细尚未产出|")
    else:
        for row in exact_audit.itertuples(index=False):
            lines.append(
                f"|{row.horizon_trading_days}|{row.model_population}|"
                f"{row.aggregation}|{row.sample_count}|"
                f"{_fmt_pct(row.net_win_rate)}|"
                f"{_fmt_pct(row.mean_net_return)}|"
                f"{_fmt_pct(row.mean_net_excess_return)}|"
            )
    lines.extend(
        [
            "",
            "## 五、真实资金模拟（AVAILABLE、质量排序、每日上限5）",
            "",
            "|期限|总收益|CAGR|最大回撤|最长水下天数|交易数|胜率|PF|费用|平均持仓|资金占用|录取率|",
            "|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|",
        ]
    )
    main = summary[
        summary["scope"].eq("AVAILABLE")
        & summary["daily_entry_limit"].eq("5")
        & summary["ranking_policy"].eq("quality")
    ].sort_values("horizon_trading_days")
    for row in main.itertuples(index=False):
        lines.append(
            f"|{row.horizon_trading_days}|{_fmt_pct(row.total_return)}|"
            f"{_fmt_pct(row.cagr)}|{_fmt_pct(row.max_drawdown)}|"
            f"{row.longest_underwater_days:.1f}|{row.closed_trades}|"
            f"{_fmt_pct(row.closed_win_rate)}|"
            f"{'NA' if pd.isna(row.profit_factor) else f'{row.profit_factor:.2f}'}|"
            f"¥{row.total_fees:,.0f}|{row.average_positions:.2f}|"
            f"{_fmt_pct(row.average_capital_utilization)}|"
            f"{_fmt_pct(row.candidate_acceptance_rate)}|"
        )
    audit_main = summary[
        summary["scope"].eq("AUDIT")
        & summary["daily_entry_limit"].eq("5")
        & summary["ranking_policy"].eq("quality")
    ].sort_values("horizon_trading_days")
    lines.extend(
        [
            "",
            "### 冻结配置的资金 AUDIT（质量排序、每日上限5）",
            "",
            "|期限|总收益|CAGR|最大回撤|交易数|胜率|PF|费用|资金占用|录取率|",
            "|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|",
        ]
    )
    for row in audit_main.itertuples(index=False):
        lines.append(
            f"|{row.horizon_trading_days}|{_fmt_pct(row.total_return)}|"
            f"{_fmt_pct(row.cagr)}|{_fmt_pct(row.max_drawdown)}|"
            f"{row.closed_trades}|{_fmt_pct(row.closed_win_rate)}|"
            f"{'NA' if pd.isna(row.profit_factor) else f'{row.profit_factor:.2f}'}|"
            f"¥{row.total_fees:,.0f}|{_fmt_pct(row.average_capital_utilization)}|"
            f"{_fmt_pct(row.candidate_acceptance_rate)}|"
        )
    lines.append("")
    if audit_main.empty:
        lines.append("- 本次运行未执行资金AUDIT。")
    else:
        for row in audit_main.itertuples(index=False):
            lines.append(
                f"- **{row.horizon_trading_days}日资金AUDIT**："
                f"{_portfolio_quality_statement(pd.Series(row._asdict()))}"
            )
        audit_start = pd.to_datetime(audit_main["start_at"]).min()
        audit_end = pd.to_datetime(audit_main["end_at"]).max()
        lines.append(
            f"- 资金AUDIT实际区间：`{audit_start.isoformat()}` 至 "
            f"`{audit_end.isoformat()}`；季度收益包含跨季持仓盯市损益。"
        )
    extrema = _audit_period_extrema(period_metrics)
    lines.extend(
        [
            "",
            "### 资金AUDIT年度、季度与行情分期",
            "",
            "|期限|维度|收益最高分期|该期收益|交易数|平仓胜率|收益最低分期|该期收益|",
            "|---:|---|---|---:|---:|---:|---|---:|",
        ]
    )
    if extrema.empty:
        lines.append("|—|—|—|—|—|—|—|资金AUDIT分期结果缺失|")
    else:
        period_labels = {"YEAR": "年度", "QUARTER": "季度", "REGIME": "行情"}
        for row in extrema.sort_values(
            ["horizon_trading_days", "period_type"],
            kind="stable",
        ).itertuples(index=False):
            lines.append(
                f"|{row.horizon_trading_days}|"
                f"{period_labels.get(row.period_type, row.period_type)}|"
                f"{row.best_period_id}|{_fmt_pct(row.best_return)}|"
                f"{row.best_closed_trades}|{_fmt_pct(row.best_closed_win_rate)}|"
                f"{row.worst_period_id}|{_fmt_pct(row.worst_return)}|"
            )
    lines.extend(
        [
            "",
            (
                "完整的每日容量限制、年度/季度/行情阶段、30分钟净值与回撤、"
                f"{random_output_text}见同目录 CSV/Parquet；Excel 的 `Charts` "
                "工作表内嵌5/30/60/90日归一净值与回撤图。"
            ),
            "",
            "## 六、与日线基准直接对照（每日上限5）",
            "",
            "|期限|30分钟总收益|日线总收益|30分钟CAGR|日线CAGR|30分钟MDD|日线MDD|",
            "|---:|---:|---:|---:|---:|---:|---:|",
        ]
    )
    for row in baseline.sort_values("horizon_trading_days").itertuples(index=False):
        lines.append(
            f"|{row.horizon_trading_days}|{_fmt_pct(row.portfolio_total_return_30m)}|"
            f"{_fmt_pct(row.portfolio_total_return_daily)}|"
            f"{_fmt_pct(row.portfolio_cagr_30m)}|{_fmt_pct(row.portfolio_cagr_daily)}|"
            f"{_fmt_pct(row.portfolio_mdd_30m)}|{_fmt_pct(row.portfolio_mdd_daily)}|"
        )
    selected_models = set(locked["model_code"].astype(str).str.split(",").explode())
    preferred = [
        code for code in ("S0006", "S0016", "S0000") if code in selected_models
    ]
    reentered = [code for code in ("S0008", "S0013") if code in selected_models]
    lines.extend(
        [
            "",
            "> 日线和30分钟样本区间不相同；上表是合同基准对照，不把差值解释成纯周期因果效应。",
            "",
            "## 七、逐项回答",
            "",
        ]
    )
    for horizon in (30, 60):
        row = main[main["horizon_trading_days"].eq(horizon)]
        if len(row):
            item = row.iloc[0]
            audit_row = audit_main[audit_main["horizon_trading_days"].eq(horizon)]
            audit_item = audit_row.iloc[0] if len(audit_row) else None
            lines.append(
                f"- **{horizon}日稳定性**：" f"{_stability_statement(item, audit_item)}"
            )
    five = main[main["horizon_trading_days"].eq(5)]
    if len(five):
        item = five.iloc[0]
        audit_five = audit_main[audit_main["horizon_trading_days"].eq(5)]
        audit_five_text = (
            _portfolio_quality_statement(audit_five.iloc[0])
            if len(audit_five)
            else "资金AUDIT未执行。"
        )
        lines.append(
            f"- **5日样本外表现**：日线AUDIT胜率47.61%；"
            f"30分钟{audit_five_text} AVAILABLE总收益为 "
            f"`{_fmt_pct(item.total_return)}`。"
        )
    audit_ninety = audit_main[audit_main["horizon_trading_days"].eq(90)]
    audit_ninety_text = (
        _portfolio_quality_statement(audit_ninety.iloc[0])
        if len(audit_ninety)
        else "资金AUDIT未执行。"
    )
    ninety_periods = (
        extrema[extrema["horizon_trading_days"].eq(90)]
        if not extrema.empty
        else extrema
    )
    ninety_best = (
        "；".join(
            f"{row.period_type}最高为{row.best_period_id}"
            f"（{_fmt_pct(row.best_return)}）"
            for row in ninety_periods.itertuples(index=False)
        )
        if not ninety_periods.empty
        else "资金AUDIT分期结果缺失"
    )
    lines.extend(
        [
            (
                f"- **90日高胜率来源**：本轮证据等级为 `{grade}`，资金曲线实际始于"
                f"`{minimum.year}`年；30分钟{audit_ninety_text}"
                f" 分期中{ninety_best}。来源判断只适用于上述实际样本区间。"
            ),
            (
                f"- **模型优先级**：锁定冠军中 S0006/S0016/S0000 命中为"
                f"`{preferred or '无'}`。"
            ),
            (
                f"- **S0008/S0013重新进入**：锁定冠军中命中为 `{reentered or '无'}`；"
                "这只表示TRAIN+VALIDATION选择结果。"
            ),
            (
                "- **费用覆盖**：表内净值已逐笔扣买卖各0.02%，累计费用和换手已输出；"
                "更高信号频率是否值得，以净CAGR、MDD和录取率共同判断。"
            ),
            (
                "- **新增信息还是拆密**：用候选事件数/Union去重信号数、同股同揭示模型数、"
                "录取率及随机排序区间共同审计；高重复且收益无改善更接近“拆得更密”。"
            ),
            "",
            "## 八、产物与复现",
            "",
            f"- 结果目录：`{(root / 'portfolio').resolve()}`",
            "- 复现命令见 `portfolio/reproduce_command.txt`；输入与输出SHA256见 `portfolio/manifest.json`。",
            (
                "- 复现前须核对稳定仓库脚本的SHA256与 manifest 中"
                f"`portfolio_logic_sha256={portfolio_logic_sha256}`一致。"
            ),
            "- 数据事实、研究假设、样本内锁定、AUDIT揭示与资金模拟在本报告中分区呈现。",
            "",
        ]
    )
    return "\n".join(lines)


def _read_json_object(path: Path, label: str) -> dict[str, Any]:
    if not path.is_file():
        raise PortfolioContractError(f"required {label} is missing: {path}")
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError) as exc:
        raise PortfolioContractError(f"{label} is unreadable: {path}") from exc
    if not isinstance(payload, Mapping):
        raise PortfolioContractError(f"{label} must be a JSON object")
    return dict(payload)


def _load_feature_summary(root: Path) -> dict[str, Any]:
    payload = _read_json_object(
        root / "features" / "summary.json",
        "features/summary.json",
    )
    if payload.get("study_id") != STUDY_ID:
        raise PortfolioContractError("features/summary.json study_id mismatch")
    return payload


def validate_feature_inputs(root: Path) -> dict[str, Any]:
    """Cross-check the frozen feature manifest, replay, snapshot, and Parquet."""

    feature_manifest = _read_json_object(
        root / "features" / "manifest.json",
        "features/manifest.json",
    )
    feature_summary = _load_feature_summary(root)
    snapshot_manifest = _read_json_object(
        root / "snapshot" / "manifest.json",
        "snapshot/manifest.json",
    )
    replay_manifest = _read_json_object(
        root / "replay" / "manifest.json",
        "replay/manifest.json",
    )
    for label, payload in (
        ("features/manifest.json", feature_manifest),
        ("snapshot/manifest.json", snapshot_manifest),
        ("replay/manifest.json", replay_manifest),
    ):
        if payload.get("study_id") != STUDY_ID:
            raise PortfolioContractError(f"{label} study_id mismatch")
    if feature_manifest.get("summary") != feature_summary:
        raise PortfolioContractError(
            "features/manifest.json summary disagrees with features/summary.json"
        )
    snapshot_id = feature_manifest.get("snapshot_id")
    if (
        not isinstance(snapshot_id, str)
        or not snapshot_id.startswith("sha256:")
        or snapshot_manifest.get("snapshot_id") != snapshot_id
        or replay_manifest.get("snapshot_id") != snapshot_id
    ):
        raise PortfolioContractError("features snapshot_id lineage mismatch")
    signal_set_id = feature_manifest.get("signal_set_id")
    if (
        not isinstance(signal_set_id, str)
        or len(signal_set_id) != 71
        or not signal_set_id.startswith("sha256:")
        or any(
            character not in "0123456789abcdef"
            for character in signal_set_id.removeprefix("sha256:")
        )
        or replay_manifest.get("signal_set_id") != signal_set_id
    ):
        raise PortfolioContractError("features signal_set_id lineage mismatch")

    output = feature_summary.get("output")
    if not isinstance(output, Mapping):
        raise PortfolioContractError("features summary.output must be an object")
    candidate_path = root / "features" / "candidate_events.parquet"
    if Path(str(output.get("path", ""))).resolve() != candidate_path.resolve():
        raise PortfolioContractError("features candidate output path mismatch")
    candidate_identity = _verify_declared_snapshot_file(
        path=candidate_path,
        logical_path="features/candidate_events.parquet",
        metadata=output,
        label="features candidate_events.parquet",
    )
    declared_rows = feature_summary.get("candidate_event_rows")
    if (
        not isinstance(declared_rows, int)
        or isinstance(declared_rows, bool)
        or declared_rows < 0
    ):
        raise PortfolioContractError("features candidate_event_rows is invalid")
    actual_rows = pq.ParquetFile(candidate_path).metadata.num_rows
    if actual_rows != declared_rows:
        raise PortfolioContractError(
            "features candidate_events Parquet row count mismatch"
        )
    return {
        "status": "VERIFIED",
        "study_id": STUDY_ID,
        "snapshot_id": snapshot_id,
        "signal_set_id": signal_set_id,
        "candidate_event_rows": actual_rows,
        "candidate_events": candidate_identity,
    }


def _load_index_benchmark(root: Path) -> dict[str, Any]:
    payload = _read_json_object(
        root / "audit" / "study_config.json",
        "audit/study_config.json",
    )
    if payload.get("study_id") != STUDY_ID:
        raise PortfolioContractError("audit/study_config.json study_id mismatch")
    source = payload.get("index_source")
    if not isinstance(source, Mapping):
        raise PortfolioContractError(
            "audit/study_config.json index_source must be an object"
        )
    source_kind = str(source.get("source_kind", "")).strip()
    source_code = str(source.get("source_code", "")).strip()
    source_name = str(source.get("source_name", "")).strip()
    if not source_name:
        raise PortfolioContractError("index_source source_name is missing")
    if source_kind == "SHANGHAI_COMPOSITE_ETF_PROXY":
        if source_code != "510980":
            raise PortfolioContractError(
                "Shanghai Composite ETF proxy must use source_code=510980"
            )
        benchmark_label = "510980上证综合ETF代理"
        is_proxy = True
    elif source_kind == "SHANGHAI_COMPOSITE":
        if source_code != "000001":
            raise PortfolioContractError(
                "Shanghai Composite source must use source_code=000001"
            )
        benchmark_label = "000001上证指数"
        is_proxy = False
    else:
        raise PortfolioContractError(
            f"unsupported index_source source_kind: {source_kind or '<missing>'}"
        )
    return {
        "source_kind": source_kind,
        "source_code": source_code,
        "source_name": source_name,
        "benchmark_label": benchmark_label,
        "is_proxy": is_proxy,
    }


def _verify_declared_snapshot_file(
    *,
    path: Path,
    logical_path: str,
    metadata: Mapping[str, Any],
    label: str,
) -> dict[str, Any]:
    if not path.is_file():
        raise PortfolioContractError(f"{label} is missing: {path}")
    expected_size = metadata.get("file_size")
    if (
        not isinstance(expected_size, int)
        or isinstance(expected_size, bool)
        or expected_size < 0
    ):
        raise PortfolioContractError(f"{label} has invalid declared file_size")
    actual_size = path.stat().st_size
    if actual_size != expected_size:
        raise PortfolioContractError(
            f"{label} size mismatch: expected {expected_size}, got {actual_size}"
        )
    expected_sha256 = metadata.get("file_sha256")
    if not isinstance(expected_sha256, str) or len(expected_sha256) != 64:
        raise PortfolioContractError(f"{label} has invalid declared file_sha256")
    actual_sha256 = sha256_file(path)
    if actual_sha256 != expected_sha256.lower():
        raise PortfolioContractError(f"{label} SHA256 mismatch")
    return {
        "logical_path": logical_path,
        "file_size": actual_size,
        "sha256": actual_sha256,
    }


def validate_snapshot_inputs(
    root: Path,
    *,
    index_benchmark: Mapping[str, Any],
) -> dict[str, Any]:
    """Verify every immutable index/bar file declared by the snapshot manifest."""

    manifest_path = root / "snapshot" / "manifest.json"
    manifest = _read_json_object(manifest_path, "snapshot/manifest.json")
    if manifest.get("study_id") != STUDY_ID:
        raise PortfolioContractError("snapshot/manifest.json study_id mismatch")
    snapshot_id = manifest.get("snapshot_id")
    if not isinstance(snapshot_id, str) or not snapshot_id.startswith("sha256:"):
        raise PortfolioContractError("snapshot/manifest.json snapshot_id is invalid")

    index_meta = manifest.get("index")
    if not isinstance(index_meta, Mapping):
        raise PortfolioContractError("snapshot manifest index must be an object")
    index_logical_path = "snapshot/index_day.parquet"
    if index_meta.get("logical_path") != index_logical_path:
        raise PortfolioContractError(
            "snapshot index logical_path must be snapshot/index_day.parquet"
        )
    for field in ("source_kind", "source_code", "source_name"):
        if index_meta.get(field) != index_benchmark.get(field):
            raise PortfolioContractError(
                f"snapshot index {field} disagrees with audit/study_config.json"
            )
    index_path = root / "snapshot" / "index_day.parquet"
    if not index_path.resolve().is_relative_to(root.resolve()):
        raise PortfolioContractError("snapshot index path escapes the study root")
    index_identity = _verify_declared_snapshot_file(
        path=index_path,
        logical_path=index_logical_path,
        metadata=index_meta,
        label="snapshot index_day.parquet",
    )

    code_files = manifest.get("code_files")
    if not isinstance(code_files, list) or not code_files:
        raise PortfolioContractError(
            "snapshot manifest code_files must be a non-empty list"
        )
    bars_dir = root / "snapshot" / "bars"
    if not bars_dir.resolve().is_relative_to(root.resolve()):
        raise PortfolioContractError("snapshot bars directory escapes the study root")
    declared_codes: set[str] = set()
    declared_logical_paths: set[str] = set()
    bar_identities: list[dict[str, Any]] = []
    bar_rows = 0
    for offset, item in enumerate(code_files):
        if not isinstance(item, Mapping):
            raise PortfolioContractError(
                f"snapshot code_files[{offset}] must be an object"
            )
        code = item.get("code")
        if not isinstance(code, str) or len(code) != 6 or not code.isdigit():
            raise PortfolioContractError(
                f"snapshot code_files[{offset}] has invalid code"
            )
        if code in declared_codes:
            raise PortfolioContractError(
                f"snapshot manifest has duplicate code: {code}"
            )
        declared_codes.add(code)
        logical_path = f"snapshot/bars/{code}.parquet"
        for path_field in ("logical_path", "path"):
            supplied_path = item.get(path_field)
            if supplied_path is not None and supplied_path != logical_path:
                raise PortfolioContractError(
                    f"snapshot bars/{code}.parquet {path_field} mismatch"
                )
        path = root / "snapshot" / "bars" / f"{code}.parquet"
        if (
            not path.resolve().is_relative_to(root.resolve())
            or path.resolve().parent != bars_dir.resolve()
        ):
            raise PortfolioContractError(
                f"snapshot bars/{code}.parquet path escapes the bars directory"
            )
        declared_logical_paths.add(logical_path)
        identity = _verify_declared_snapshot_file(
            path=path,
            logical_path=logical_path,
            metadata=item,
            label=f"snapshot bars/{code}.parquet",
        )
        identity["code"] = code
        rows = item.get("rows")
        if not isinstance(rows, int) or isinstance(rows, bool) or rows < 0:
            raise PortfolioContractError(
                f"snapshot bars/{code}.parquet has invalid declared rows"
            )
        actual_rows = pq.ParquetFile(path).metadata.num_rows
        if actual_rows != rows:
            raise PortfolioContractError(
                f"snapshot bars/{code}.parquet Parquet row count mismatch"
            )
        identity["rows"] = actual_rows
        bar_identities.append(identity)
        bar_rows += rows

    actual_logical_paths = {
        path.relative_to(root).as_posix()
        for path in bars_dir.rglob("*.parquet")
        if path.is_file()
    }
    if actual_logical_paths != declared_logical_paths:
        missing = len(declared_logical_paths - actual_logical_paths)
        extra = len(actual_logical_paths - declared_logical_paths)
        raise PortfolioContractError(
            "snapshot bars file set disagrees with manifest "
            f"(missing={missing}, extra={extra})"
        )

    verified_identities = [index_identity, *bar_identities]
    identities_sha256 = hashlib.sha256(
        _canonical_bytes(verified_identities)
    ).hexdigest()
    return {
        "status": "VERIFIED",
        "study_id": STUDY_ID,
        "snapshot_id": snapshot_id,
        "all_declared_files_verified": True,
        "verified_file_count": len(verified_identities),
        "verified_bar_file_count": len(bar_identities),
        "verified_bar_rows": bar_rows,
        "verified_total_bytes": sum(
            int(identity["file_size"]) for identity in verified_identities
        ),
        "verified_identities_sha256": identities_sha256,
        "index_source": {
            field: index_benchmark[field]
            for field in ("source_kind", "source_code", "source_name")
        },
    }


def _manifest_entry(
    path: Path, root: Path, *, rows: int | None = None
) -> dict[str, Any]:
    return {
        "logical_path": path.relative_to(root).as_posix(),
        "rows": rows,
        "file_size": path.stat().st_size,
        "sha256": sha256_file(path),
    }


def _required_input_identity(path: Path, root: Path) -> dict[str, Any]:
    if not path.is_file():
        raise PortfolioContractError(f"required portfolio input is missing: {path}")
    return {
        "logical_path": path.relative_to(root).as_posix(),
        "file_size": path.stat().st_size,
        "sha256": sha256_file(path),
    }


def _run_input_paths(root: Path) -> dict[str, Path]:
    return {
        "candidate_events": root / "features" / "candidate_events.parquet",
        "features_manifest": root / "features" / "manifest.json",
        "feature_summary": root / "features" / "summary.json",
        "replay_manifest": root / "replay" / "manifest.json",
        "market_segments": root / "features" / "market_segments.csv",
        "locked_config": root / "matrix" / "locked_config.json",
        "lock_manifest": root / "matrix" / "lock_manifest.json",
        "development_manifest": root / "matrix" / "development_manifest.json",
        "development_lock_candidates": (
            root / "matrix" / "development_lock_candidates.parquet"
        ),
        "reveal_manifest": root / "matrix" / "reveal_manifest.json",
        "reveal_matrix": root / "matrix" / "reveal_matrix.parquet",
        "reveal_summary": root / "matrix" / "reveal_summary.csv",
        "reveal_locked_detailed": (root / "matrix" / "reveal_locked_detailed.parquet"),
        "reveal_locked_group_detail": (
            root / "matrix" / "reveal_locked_group_detail.parquet"
        ),
        "study_config": root / "audit" / "study_config.json",
        "snapshot_manifest": root / "snapshot" / "manifest.json",
        "index_day": root / "snapshot" / "index_day.parquet",
    }


def _current_run_input_identities(root: Path) -> dict[str, dict[str, Any]]:
    return {
        name: _required_input_identity(path, root)
        for name, path in _run_input_paths(root).items()
    }


def _validate_frozen_run_inputs(
    root: Path,
    run_contract: Mapping[str, Any],
) -> dict[str, dict[str, Any]]:
    frozen = run_contract.get("input_identities")
    if not isinstance(frozen, Mapping):
        raise PortfolioContractError("frozen run_contract input_identities is missing")
    current = _current_run_input_identities(root)
    if set(frozen) != set(current):
        raise PortfolioContractError("frozen run_contract input set mismatch")
    for name, identity in current.items():
        if frozen.get(name) != identity:
            raise PortfolioContractError(
                f"frozen run_contract external input drift: {name}"
            )
    return current


def _checkpoint_key(
    *,
    selection: LockedSelection,
    scope: str,
    daily_entry_limit: int | None,
    ranking_policy: str,
    random_seed: int | None,
) -> str:
    seed = "n" if random_seed is None else f"{random_seed:03d}"
    cap = "u" if daily_entry_limit is None else str(daily_entry_limit)
    try:
        scope_token = CHECKPOINT_SCOPE_TOKENS[scope]
    except KeyError as exc:
        raise PortfolioContractError(f"unsupported checkpoint scope: {scope}") from exc
    return f"h{selection.horizon}_{scope_token}_c{cap}_" f"{ranking_policy[0]}_s{seed}"


def _load_quality_checkpoint(path: Path, *, run_id: str) -> SimulationResult | None:
    complete = path / "complete.json"
    if not complete.is_file():
        return None
    try:
        metadata = json.loads(complete.read_text(encoding="utf-8"))
        if metadata.get("run_id") != run_id:
            return None
        equity_path = path / "equity.parquet"
        trades_path = path / "trades.parquet"
        if metadata.get("equity_sha256") != sha256_file(equity_path) or metadata.get(
            "trades_sha256"
        ) != sha256_file(trades_path):
            return None
        summary = dict(metadata["summary"])
        for column in ("start_at", "end_at"):
            if summary.get(column):
                summary[column] = pd.Timestamp(summary[column])
        return SimulationResult(
            summary=summary,
            equity=pd.read_parquet(equity_path),
            trades=pd.read_parquet(trades_path),
            decisions=pd.DataFrame(),
        )
    except (FileNotFoundError, KeyError, OSError, ValueError):
        return None


def _save_quality_checkpoint(
    path: Path, *, run_id: str, result: SimulationResult
) -> None:
    path.mkdir(parents=True, exist_ok=True)
    equity_path = path / "equity.parquet"
    trades_path = path / "trades.parquet"
    _atomic_parquet(result.equity, equity_path)
    _atomic_parquet(result.trades, trades_path)
    _atomic_json(
        path / "complete.json",
        {
            "run_id": run_id,
            "summary": result.summary,
            "equity_sha256": sha256_file(equity_path),
            "trades_sha256": sha256_file(trades_path),
        },
    )


def _load_random_checkpoint(path: Path, *, run_id: str) -> dict[str, Any] | None:
    if not path.is_file():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if payload.get("run_id") != run_id:
            return None
        return dict(payload["summary"])
    except (KeyError, OSError, ValueError):
        return None


def _save_random_checkpoint(
    path: Path, *, run_id: str, summary: Mapping[str, Any]
) -> None:
    _atomic_json(path, {"run_id": run_id, "summary": dict(summary)})


def _completed_result(root: Path, *, run_id: str) -> dict[str, Any] | None:
    output = root / "portfolio"
    manifest_path = output / "manifest.json"
    if not manifest_path.is_file():
        return None
    manifest = _read_json_object(manifest_path, "portfolio/manifest.json")
    if manifest.get("status") != "COMPLETE" or manifest.get("run_id") != run_id:
        return None
    if manifest.get("study_id") != STUDY_ID:
        raise PortfolioContractError("completed portfolio study_id mismatch")
    if manifest.get("portfolio_logic_sha256") != sha256_file(Path(__file__).resolve()):
        raise PortfolioContractError("completed portfolio logic SHA256 mismatch")
    outputs = manifest.get("outputs")
    if not isinstance(outputs, list) or not outputs:
        raise PortfolioContractError(
            "completed portfolio outputs must be a non-empty list"
        )
    logical_paths: list[str] = []
    for offset, item in enumerate(outputs):
        if not isinstance(item, Mapping):
            raise PortfolioContractError(
                f"completed portfolio outputs[{offset}] must be an object"
            )
        logical_path = item.get("logical_path")
        if not isinstance(logical_path, str) or not logical_path:
            raise PortfolioContractError(
                f"completed portfolio outputs[{offset}] logical_path is invalid"
            )
        logical_paths.append(logical_path)
    if len(logical_paths) != len(set(logical_paths)):
        raise PortfolioContractError(
            "completed portfolio outputs contain duplicate logical_path values"
        )
    actual_outputs = set(logical_paths)
    if actual_outputs != REQUIRED_LOGICAL_OUTPUTS:
        raise PortfolioContractError(
            "completed portfolio required logical outputs mismatch "
            f"(missing={len(REQUIRED_LOGICAL_OUTPUTS - actual_outputs)}, "
            f"extra={len(actual_outputs - REQUIRED_LOGICAL_OUTPUTS)})"
        )
    for item in outputs:
        logical_path = str(item["logical_path"])
        path = root / logical_path
        expected_size = item.get("file_size")
        expected_sha256 = item.get("sha256")
        if (
            not path.is_file()
            or not isinstance(expected_size, int)
            or isinstance(expected_size, bool)
            or path.stat().st_size != expected_size
            or not isinstance(expected_sha256, str)
            or len(expected_sha256) != 64
            or sha256_file(path) != expected_sha256.lower()
        ):
            raise PortfolioContractError(
                f"completed portfolio output identity mismatch: {logical_path}"
            )
    try:
        quality_portfolios = int(manifest["quality_portfolios"])
        random_portfolios = int(manifest["random_portfolios"])
    except (KeyError, TypeError, ValueError) as exc:
        raise PortfolioContractError("completed portfolio counts are invalid") from exc
    return {
        "study_id": STUDY_ID,
        "run_id": run_id,
        "root": str(root),
        "portfolio_dir": str(output),
        "quality_portfolios": quality_portfolios,
        "random_portfolios": random_portfolios,
        "report": str(output / "report.md"),
        "workbook": str(output / "clx_30m_portfolio_report.xlsx"),
        "manifest": str(manifest_path),
        "reused": True,
    }


def run_portfolios(
    *,
    root: Path,
    random_seeds: int = 100,
    include_audit_scope: bool = True,
) -> dict[str, Any]:
    root = root.resolve()
    if random_seeds < 0:
        raise PortfolioContractError("random_seeds must be non-negative")
    event_path = root / "features" / "candidate_events.parquet"
    lock_path = root / "matrix" / "locked_config.json"
    if not event_path.is_file():
        raise PortfolioContractError(f"candidate events are missing: {event_path}")
    selections = load_locked_selections(lock_path)
    reveal_artifacts = validate_reveal_inputs(root, lock_path)
    reveal_path, reveal_matrix_sha256 = reveal_artifacts["matrix"]
    reveal_summary_path, reveal_summary_sha256 = reveal_artifacts["summary"]
    reveal_detailed_path, reveal_detailed_sha256 = reveal_artifacts["locked_detailed"]
    reveal_group_path, reveal_group_sha256 = reveal_artifacts["group_detail"]
    _reveal_manifest_path, reveal_manifest_sha256 = reveal_artifacts["manifest"]
    event_sha256 = sha256_file(event_path)
    lock_sha256 = sha256_file(lock_path)
    snapshot_manifest_path = root / "snapshot" / "manifest.json"
    segment_path = root / "features" / "market_segments.csv"
    feature_summary_path = root / "features" / "summary.json"
    study_config_path = root / "audit" / "study_config.json"
    snapshot_manifest_identity = _required_input_identity(
        snapshot_manifest_path,
        root,
    )
    market_segments_identity = _required_input_identity(segment_path, root)
    feature_summary_identity = _required_input_identity(feature_summary_path, root)
    study_config_identity = _required_input_identity(study_config_path, root)
    feature_verification = validate_feature_inputs(root)
    feature_summary = _load_feature_summary(root)
    index_benchmark = _load_index_benchmark(root)
    snapshot_verification = validate_snapshot_inputs(
        root,
        index_benchmark=index_benchmark,
    )
    input_identities = _current_run_input_identities(root)
    run_contract = {
        "study_id": STUDY_ID,
        "contract_version": PORTFOLIO_CONTRACT_VERSION,
        "portfolio_logic_sha256": sha256_file(Path(__file__).resolve()),
        "candidate_events_sha256": event_sha256,
        "locked_config_sha256": lock_sha256,
        "reveal_manifest_sha256": reveal_manifest_sha256,
        "reveal_matrix_sha256": reveal_matrix_sha256,
        "reveal_summary_sha256": reveal_summary_sha256,
        "reveal_locked_detailed_sha256": reveal_detailed_sha256,
        "reveal_locked_group_detail_sha256": reveal_group_sha256,
        "snapshot_manifest": snapshot_manifest_identity,
        "market_segments": market_segments_identity,
        "feature_summary": feature_summary_identity,
        "study_config": study_config_identity,
        "input_identities": input_identities,
        "feature_verification": feature_verification,
        "snapshot_verification": snapshot_verification,
        "random_seed_count": random_seeds,
        "include_audit_scope": include_audit_scope,
        "checkpoint_scope_tokens": CHECKPOINT_SCOPE_TOKENS,
    }
    run_id = "sha256:" + hashlib.sha256(_canonical_bytes(run_contract)).hexdigest()
    reusable = _completed_result(root, run_id=run_id)
    if reusable is not None:
        print(json.dumps(reusable, ensure_ascii=False), flush=True)
        return reusable
    output = root / "portfolio"
    checkpoint_root = output / "ckpt" / run_id.removeprefix("sha256:")[:16]
    event_columns = [
        "signal_fact_id",
        "union_signal_id",
        "code",
        "model_code",
        "reveal_at",
        "entry_at",
        "qfq_entry_open",
        "entry_executable",
        "entry_status",
        "concurrent_trigger_mask",
        "concurrent_trigger_count",
        "filter_pass_mask",
        "same_code_reveal_model_count",
        "same_reveal_event_count",
        "amount_median_20d",
        "raw_entry_gap",
        "market_regime",
        "split_id",
    ]
    for horizon in HORIZONS:
        event_columns.extend(
            (
                f"h{horizon}_status",
                f"h{horizon}_exit_at",
                f"h{horizon}_gross_return",
                f"h{horizon}_split_boundary_status",
            )
        )
    locked_models = sorted(
        {
            model
            for selection in selections
            for model in selection.model_codes
            if model not in {"ALL", "*", "UNION"}
        }
    )
    wildcard_model = any(
        set(selection.model_codes) & {"ALL", "*", "UNION"} for selection in selections
    )
    events = pd.read_parquet(
        event_path,
        columns=event_columns,
        filters=(None if wildcard_model else [("model_code", "in", locked_models)]),
    )
    scopes = ["AVAILABLE"]
    if (
        include_audit_scope
        and "split_id" in events
        and events["split_id"].eq("AUDIT").any()
    ):
        scopes.append("AUDIT")
    selected: dict[tuple[str, str], pd.DataFrame] = {}
    for selection in selections:
        available = select_locked_candidates(events, selection, scope="AVAILABLE")
        selected[(selection.selection_id, "AVAILABLE")] = available
        if "AUDIT" in scopes:
            selected[(selection.selection_id, "AUDIT")] = select_locked_candidates(
                events,
                selection,
                scope="AUDIT",
            )
    available_frames = [
        selected[(selection.selection_id, "AVAILABLE")] for selection in selections
    ]
    codes = {
        str(code)
        for frame in available_frames
        for code in frame.get("code", pd.Series(dtype=str)).astype(str).unique()
    }
    marks = MarkStore(root)
    marks.load(codes)
    scope_clocks = {"AVAILABLE": build_simulation_clock(root, available_frames, marks)}
    if "AUDIT" in scopes:
        audit_frames = [
            selected[(selection.selection_id, "AUDIT")] for selection in selections
        ]
        if any(not frame.empty for frame in audit_frames):
            scope_clocks["AUDIT"] = build_simulation_clock(root, audit_frames, marks)
        else:
            scope_clocks["AUDIT"] = scope_clocks["AVAILABLE"]
    segments = pd.read_csv(segment_path, encoding="utf-8-sig")
    summaries: list[dict[str, Any]] = []
    random_summaries: list[dict[str, Any]] = []
    quality_equity: list[pd.DataFrame] = []
    quality_trades: list[pd.DataFrame] = []
    periods: list[pd.DataFrame] = []
    for selection in selections:
        for scope in scopes:
            frame = selected[(selection.selection_id, scope)]
            quality_entries: dict[int, list[dict[str, Any]]] | None = None
            for daily_limit in DAILY_ENTRY_LIMITS:
                checkpoint_path = (
                    checkpoint_root
                    / "quality"
                    / _checkpoint_key(
                        selection=selection,
                        scope=scope,
                        daily_entry_limit=daily_limit,
                        ranking_policy="quality",
                        random_seed=None,
                    )
                )
                quality = _load_quality_checkpoint(checkpoint_path, run_id=run_id)
                if quality is None:
                    if quality_entries is None:
                        quality_entries = prepare_ordered_entries(
                            frame,
                            ranking_policy="quality",
                        )
                    quality = simulate_portfolio(
                        frame,
                        selection=selection,
                        scope=scope,
                        clock=scope_clocks[scope],
                        marks=marks,
                        daily_entry_limit=daily_limit,
                        ranking_policy="quality",
                        ordered_entries=quality_entries,
                    )
                    _save_quality_checkpoint(
                        checkpoint_path,
                        run_id=run_id,
                        result=quality,
                    )
                summaries.append(quality.summary)
                quality_equity.append(quality.equity)
                quality_trades.append(quality.trades)
                periods.append(
                    build_period_metrics(quality.equity, quality.trades, segments)
                )
            # Random ordering is a capacity sensitivity on the full AVAILABLE
            # sequence. The order is independent of the daily cap, so one
            # SHA ordering is reused by all six accounts for that seed.
            if scope == "AVAILABLE":
                for seed in range(random_seeds):
                    pending: list[tuple[int | None, Path]] = []
                    seed_summaries: dict[str, dict[str, Any]] = {}
                    for daily_limit in DAILY_ENTRY_LIMITS:
                        checkpoint_path = (
                            checkpoint_root
                            / "random"
                            / (
                                _checkpoint_key(
                                    selection=selection,
                                    scope=scope,
                                    daily_entry_limit=daily_limit,
                                    ranking_policy="sha_random",
                                    random_seed=seed,
                                )
                                + ".json"
                            )
                        )
                        cached = _load_random_checkpoint(checkpoint_path, run_id=run_id)
                        if cached is None:
                            pending.append((daily_limit, checkpoint_path))
                        else:
                            seed_summaries[_limit_label(daily_limit)] = cached
                    if pending:
                        random_entries = prepare_ordered_entries(
                            frame,
                            ranking_policy="sha_random",
                            random_seed=seed,
                        )
                        for daily_limit, checkpoint_path in pending:
                            random = simulate_portfolio(
                                frame,
                                selection=selection,
                                scope=scope,
                                clock=scope_clocks[scope],
                                marks=marks,
                                daily_entry_limit=daily_limit,
                                ranking_policy="sha_random",
                                random_seed=seed,
                                ordered_entries=random_entries,
                            )
                            _save_random_checkpoint(
                                checkpoint_path,
                                run_id=run_id,
                                summary=random.summary,
                            )
                            seed_summaries[_limit_label(daily_limit)] = random.summary
                    random_summaries.extend(
                        seed_summaries[_limit_label(daily_limit)]
                        for daily_limit in DAILY_ENTRY_LIMITS
                    )
    summary_frame = pd.DataFrame(summaries)
    random_frame = pd.DataFrame(random_summaries)
    random_summary = summarise_random_runs(random_frame)
    equity_frame = pd.concat(quality_equity, ignore_index=True)
    trade_frame = (
        pd.concat(quality_trades, ignore_index=True)
        if any(not frame.empty for frame in quality_trades)
        else pd.DataFrame()
    )
    decision_frame = summary_frame.loc[
        :,
        [
            "selection_id",
            "horizon_trading_days",
            "scope",
            "daily_entry_limit",
            "ranking_policy",
            "candidate_signals",
            "closed_trades",
            "candidate_acceptance_rate",
            "rejected_occupied",
            "rejected_daily_limit",
            "rejected_slots",
            "rejected_cash",
        ],
    ].copy()
    period_frame = (
        pd.concat(periods, ignore_index=True)
        if any(not frame.empty for frame in periods)
        else pd.DataFrame()
    )
    locked_frame = pd.DataFrame([selection.as_row() for selection in selections])
    baseline_frame = build_daily_baseline_comparison(summary_frame)
    daily_curve_frame = _daily_curves(equity_frame)
    chart_frame = build_chart_data(daily_curve_frame)
    reveal_path = root / "matrix" / "reveal_matrix.parquet"
    reveal_summary_path = root / "matrix" / "reveal_summary.csv"
    reveal_frame = (
        pd.read_csv(reveal_summary_path, encoding="utf-8-sig")
        if reveal_summary_path.is_file()
        else pd.DataFrame()
    )
    reveal_detailed_frame = pd.read_parquet(reveal_detailed_path)
    reveal_group_frame = pd.read_parquet(reveal_group_path)
    output = root / "portfolio"
    output.mkdir(parents=True, exist_ok=True)
    frames = {
        "portfolio_summary": summary_frame,
        "random_order_runs": random_frame,
        "random_order_sensitivity": random_summary,
        "period_metrics": period_frame,
        "equity_30m": equity_frame,
        "equity_daily": daily_curve_frame,
        "chart_curves": chart_frame,
        "trades": trade_frame,
        "decision_summary": decision_frame,
        "locked_selections": locked_frame,
        "daily_baseline_comparison": baseline_frame,
    }
    if set(frames) != set(PORTFOLIO_FRAME_NAMES):
        raise PortfolioContractError("portfolio logical output frame set drifted")
    written: list[Path] = []
    output_rows: dict[Path, int] = {}
    for name, frame in frames.items():
        parquet = output / f"{name}.parquet"
        csv = output / f"{name}.csv"
        _atomic_parquet(frame, parquet)
        _atomic_csv(frame, csv)
        written.extend((parquet, csv))
        output_rows[parquet] = len(frame)
        output_rows[csv] = len(frame)
    workbook = output / "clx_30m_portfolio_report.xlsx"
    _write_excel(
        workbook,
        summary=summary_frame,
        random_summary=random_summary,
        period_metrics=period_frame,
        daily_curves=daily_curve_frame,
        chart_data=chart_frame,
        locked=locked_frame,
        baseline=baseline_frame,
        reveal=reveal_frame,
        reveal_detailed=reveal_detailed_frame,
        reveal_groups=reveal_group_frame,
    )
    written.append(workbook)
    report = build_markdown_report(
        root=root,
        summary=summary_frame,
        random_summary=random_summary,
        period_metrics=period_frame,
        locked=locked_frame,
        baseline=baseline_frame,
        feature_summary=feature_summary,
        reveal_summary=reveal_frame,
        reveal_detailed=reveal_detailed_frame,
        index_benchmark=index_benchmark,
        portfolio_logic_sha256=str(run_contract["portfolio_logic_sha256"]),
    )
    report_path = output / "report.md"
    report_path.write_text(report, encoding="utf-8")
    written.append(report_path)
    config = {
        "study_id": STUDY_ID,
        "run_id": run_id,
        "run_contract": run_contract,
        "contract_version": PORTFOLIO_CONTRACT_VERSION,
        "filter_contract": "F1_F6_64_SUBSETS_ONLY",
        "initial_capital": INITIAL_CAPITAL,
        "max_positions": MAX_POSITIONS,
        "slot_capital_budget": SLOT_CAPITAL,
        "fee_per_side": FEE_PER_SIDE,
        "daily_entry_limits": [_limit_label(value) for value in DAILY_ENTRY_LIMITS],
        "ordering": {
            "quality": (
                "within identical entry_at: concurrent trigger count desc, "
                "same-code model count desc, crowding asc, liquidity desc, "
                "entry gap asc, candidate_id"
            ),
            "sensitivity": "SHA256(seed|candidate_id)",
            "random_seed_count": random_seeds,
        },
        "execution": {
            "candidate_read": "Parquet projection plus locked-model predicate",
            "ordering_reuse": "one ordering per selection/scope/seed shared by six caps",
            "mark_to_market": "vectorised trade slices over cached clock-aligned QFQ closes",
            "checkpoint": "immutable run-id quality Parquet and random-summary JSON",
            "completed_rerun": "verify output SHA256 and return REUSED",
        },
        "clock": "30-minute QFQ close; carry prior close through missing bars",
        "metric_formulas": {
            "one_way_turnover": "(buy_notional+sell_notional)/(2*mean_equity)",
            "two_way_turnover": "(buy_notional+sell_notional)/mean_equity",
            "profit_factor": "sum(positive_net_pnl)/abs(sum(negative_net_pnl))",
            "capital_utilization": "marked_position_value/equity",
        },
        "filter_descriptions": {
            **FILTER_DESCRIPTIONS,
            "F6": (f"{index_benchmark['benchmark_label']}" "近20个完整交易日收益≤0"),
        },
        "index_benchmark": index_benchmark,
        "feature_verification": feature_verification,
        "snapshot_verification": snapshot_verification,
        "omitted_costs": [
            "slippage",
            "stamp_duty",
            "minimum_commission",
            "100_share_rounding",
        ],
        "scopes": scopes,
        "input": {
            "snapshot_manifest": snapshot_manifest_identity,
            "market_segments": market_segments_identity,
            "feature_summary": feature_summary_identity,
            "study_config": study_config_identity,
            "all_frozen_inputs": input_identities,
        },
        "candidate_events_sha256": event_sha256,
        "locked_config_sha256": lock_sha256,
    }
    config_path = output / "portfolio_config.json"
    _atomic_json(config_path, config)
    written.append(config_path)
    command = (
        f'& "<PYTHON>" "{REPRODUCE_SCRIPT}" --root "{root}" '
        f"run --random-seeds {random_seeds}"
    )
    if not include_audit_scope:
        command += " --no-audit-scope"
    reproduce = output / "reproduce_command.txt"
    reproduce.write_text(command + "\n", encoding="utf-8")
    written.append(reproduce)
    manifest = {
        "study_id": STUDY_ID,
        "run_id": run_id,
        "status": "COMPLETE",
        "portfolio_logic_sha256": run_contract["portfolio_logic_sha256"],
        "input": {
            "candidate_events": {
                "logical_path": "features/candidate_events.parquet",
                "file_size": event_path.stat().st_size,
                "sha256": event_sha256,
            },
            "locked_config": {
                "logical_path": "matrix/locked_config.json",
                "file_size": lock_path.stat().st_size,
                "sha256": lock_sha256,
            },
            "reveal_matrix": (
                {
                    "logical_path": "matrix/reveal_matrix.parquet",
                    "file_size": reveal_path.stat().st_size,
                    "sha256": sha256_file(reveal_path),
                }
                if reveal_path.is_file()
                else None
            ),
            "reveal_summary": (
                {
                    "logical_path": "matrix/reveal_summary.csv",
                    "file_size": reveal_summary_path.stat().st_size,
                    "sha256": sha256_file(reveal_summary_path),
                }
                if reveal_summary_path.is_file()
                else None
            ),
            "reveal_locked_detailed": {
                "logical_path": "matrix/reveal_locked_detailed.parquet",
                "file_size": reveal_detailed_path.stat().st_size,
                "sha256": reveal_detailed_sha256,
            },
            "reveal_locked_group_detail": {
                "logical_path": "matrix/reveal_locked_group_detail.parquet",
                "file_size": reveal_group_path.stat().st_size,
                "sha256": reveal_group_sha256,
            },
            "snapshot_manifest": snapshot_manifest_identity,
            "market_segments": market_segments_identity,
            "feature_summary": feature_summary_identity,
            "study_config": study_config_identity,
        },
        "index_benchmark": index_benchmark,
        "feature_verification": feature_verification,
        "snapshot_verification": snapshot_verification,
        "selection_count": len(selections),
        "loaded_locked_model_event_rows": len(events),
        "selected_candidate_rows": {
            f"h{selection.horizon}_{scope}": len(
                selected[(selection.selection_id, scope)]
            )
            for selection in selections
            for scope in scopes
        },
        "quality_portfolios": len(summary_frame),
        "random_portfolios": len(random_frame),
        "ordering_builds": {
            "quality": len(selections) * len(scopes),
            "sha_random": len(selections) * random_seeds,
            "six_daily_caps_share_each_ordering": True,
        },
        "checkpoint_root": checkpoint_root.relative_to(root).as_posix(),
        "mark_source_codes": len(codes),
        "mark_source_rows": marks.source_rows,
        "clock_rows": {
            scope: len(scope_clock) for scope, scope_clock in scope_clocks.items()
        },
        "outputs": [
            _manifest_entry(path, root, rows=output_rows.get(path)) for path in written
        ],
    }
    manifest_path = output / "manifest.json"
    _atomic_json(manifest_path, manifest)
    result = {
        "study_id": STUDY_ID,
        "run_id": run_id,
        "root": str(root),
        "portfolio_dir": str(output),
        "quality_portfolios": len(summary_frame),
        "random_portfolios": len(random_frame),
        "report": str(report_path),
        "workbook": str(workbook),
        "manifest": str(manifest_path),
        "reused": False,
    }
    print(json.dumps(result, ensure_ascii=False), flush=True)
    return result


def rebuild_report(root: Path) -> dict[str, Any]:
    resolved = root.resolve()
    output = resolved / "portfolio"
    config = _read_json_object(
        output / "portfolio_config.json",
        "portfolio/portfolio_config.json",
    )
    if config.get("study_id") != STUDY_ID:
        raise PortfolioContractError("portfolio config study_id mismatch")
    run_contract = config.get("run_contract")
    if not isinstance(run_contract, Mapping):
        raise PortfolioContractError("portfolio config run_contract is missing")
    frozen_logic_sha256 = run_contract.get("portfolio_logic_sha256")
    current_logic_sha256 = sha256_file(Path(__file__).resolve())
    if frozen_logic_sha256 != current_logic_sha256:
        raise PortfolioContractError("rebuild-report portfolio logic SHA256 mismatch")
    manifest = _read_json_object(
        output / "manifest.json",
        "portfolio/manifest.json",
    )
    run_id = config.get("run_id")
    if (
        not isinstance(run_id, str)
        or manifest.get("run_id") != run_id
        or run_contract.get("study_id") != STUDY_ID
        or "sha256:" + hashlib.sha256(_canonical_bytes(run_contract)).hexdigest()
        != run_id
    ):
        raise PortfolioContractError("rebuild-report run identity mismatch")
    if _completed_result(resolved, run_id=run_id) is None:
        raise PortfolioContractError("rebuild-report requires a COMPLETE portfolio")

    _validate_frozen_run_inputs(resolved, run_contract)
    lock_path = resolved / "matrix" / "locked_config.json"
    load_locked_selections(lock_path)
    validate_reveal_inputs(resolved, lock_path)
    feature_verification = validate_feature_inputs(resolved)
    if run_contract.get("feature_verification") != feature_verification:
        raise PortfolioContractError("rebuild-report feature verification drift")
    feature_summary = _load_feature_summary(resolved)
    index_benchmark = _load_index_benchmark(resolved)
    if config.get("index_benchmark") != index_benchmark:
        raise PortfolioContractError("rebuild-report index benchmark mismatch")
    snapshot_verification = validate_snapshot_inputs(
        resolved,
        index_benchmark=index_benchmark,
    )
    if run_contract.get("snapshot_verification") != snapshot_verification:
        raise PortfolioContractError("rebuild-report snapshot verification drift")

    summary = pd.read_parquet(output / "portfolio_summary.parquet")
    random_summary = pd.read_parquet(output / "random_order_sensitivity.parquet")
    period_metrics = pd.read_parquet(output / "period_metrics.parquet")
    locked = pd.read_parquet(output / "locked_selections.parquet")
    baseline = pd.read_parquet(output / "daily_baseline_comparison.parquet")
    reveal_summary_path = resolved / "matrix" / "reveal_summary.csv"
    reveal_summary = (
        pd.read_csv(reveal_summary_path, encoding="utf-8-sig")
        if reveal_summary_path.is_file()
        else pd.DataFrame()
    )
    reveal_detailed_path = resolved / "matrix" / "reveal_locked_detailed.parquet"
    reveal_detailed = (
        pd.read_parquet(reveal_detailed_path)
        if reveal_detailed_path.is_file()
        else pd.DataFrame()
    )
    text = build_markdown_report(
        root=resolved,
        summary=summary,
        random_summary=random_summary,
        period_metrics=period_metrics,
        locked=locked,
        baseline=baseline,
        feature_summary=feature_summary,
        reveal_summary=reveal_summary,
        reveal_detailed=reveal_detailed,
        index_benchmark=index_benchmark,
        portfolio_logic_sha256=str(frozen_logic_sha256),
    )
    path = output / "report.md"
    path.write_text(text, encoding="utf-8")
    manifest_path = output / "manifest.json"
    for item in manifest["outputs"]:
        if item["logical_path"] == "portfolio/report.md":
            item["file_size"] = path.stat().st_size
            item["sha256"] = sha256_file(path)
            break
    _atomic_json(manifest_path, manifest)
    result = {"report": str(path), "sha256": sha256_file(path)}
    print(json.dumps(result, ensure_ascii=False), flush=True)
    return result


def status(root: Path) -> dict[str, Any]:
    root = root.resolve()
    result = {
        "study_id": STUDY_ID,
        "root": str(root),
        "candidate_events": (root / "features" / "candidate_events.parquet").is_file(),
        "locked_config": (root / "matrix" / "locked_config.json").is_file(),
        "reveal_matrix": (root / "matrix" / "reveal_matrix.parquet").is_file(),
        "reveal_summary": (root / "matrix" / "reveal_summary.csv").is_file(),
        "reveal_locked_detailed": (
            root / "matrix" / "reveal_locked_detailed.parquet"
        ).is_file(),
        "reveal_locked_group_detail": (
            root / "matrix" / "reveal_locked_group_detail.parquet"
        ).is_file(),
        "portfolio_complete": (root / "portfolio" / "manifest.json").is_file(),
    }
    print(json.dumps(result, ensure_ascii=False), flush=True)
    return result


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", type=Path, default=DEFAULT_ROOT)
    commands = parser.add_subparsers(dest="command", required=True)
    run = commands.add_parser("run")
    run.add_argument("--random-seeds", type=int, default=100)
    run.add_argument("--no-audit-scope", action="store_true")
    commands.add_parser("report")
    commands.add_parser("status")
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    if args.command == "run":
        run_portfolios(
            root=args.root,
            random_seeds=args.random_seeds,
            include_audit_scope=not args.no_audit_scope,
        )
    elif args.command == "report":
        rebuild_report(args.root)
    else:
        status(args.root)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
