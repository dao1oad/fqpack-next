from __future__ import annotations

import argparse
import json
import math
import subprocess
import sys
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

from script.clx_backtest.research.clx_30m_f1_f6_matrix import (
    _development_identity,
    _reveal_identity,
    run_lock,
    write_json_atomic,
)
from script.clx_backtest.research.clx_30m_f1_f6_portfolio import (
    FEE_PER_SIDE,
    INITIAL_CAPITAL,
    REPRODUCE_SCRIPT,
    REQUIRED_LOGICAL_OUTPUTS,
    MarkStore,
    PortfolioContractError,
    _checkpoint_key,
    _completed_result,
    _parse_locked_selections,
    _portfolio_quality_statement,
    _stability_statement,
    _write_excel,
    build_simulation_clock,
    load_locked_selections,
    rebuild_report,
    run_portfolios,
    select_locked_candidates,
    sha256_file,
    simulate_portfolio,
    validate_snapshot_inputs,
)


def test_direct_script_cli_help_from_repo_root() -> None:
    repo_root = Path(__file__).resolve().parents[3]
    script_path = (
        repo_root
        / "script"
        / "clx_backtest"
        / "research"
        / "clx_30m_f1_f6_portfolio.py"
    )
    completed = subprocess.run(
        [sys.executable, str(script_path), "--help"],
        cwd=repo_root,
        capture_output=True,
        text=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr
    assert "usage:" in completed.stdout.lower()


def _locked_payload(filter_mask: int = 0) -> dict[str, object]:
    return {
        "lock_id": "sha256:fixture-lock",
        "study_id": "clx-30m-full-trigger-f1-f6-v1",
        "selections": [
            {
                "selection_id": f"h{horizon}-fixture",
                "horizon_trading_days": horizon,
                "model_code": "S0000",
                "trigger_id": "ALL",
                "trigger_selector": {
                    "kind": "ALL",
                    "value": None,
                    "name": "全部触发",
                },
                "filter_mask": filter_mask,
                "filter_names": (["F1"] if filter_mask == 1 else []),
                "development_score": 1.0,
                "train_metrics": {"sample_count": 100, "win_rate": 0.55},
                "validation_metrics": {"sample_count": 50, "win_rate": 0.54},
            }
            for horizon in (5, 30, 60, 90)
        ],
    }


def _write_lock(root: Path, filter_mask: int = 0) -> None:
    path = root / "matrix" / "locked_config.json"
    path.parent.mkdir(parents=True)
    path.write_text(
        json.dumps(_locked_payload(filter_mask), ensure_ascii=False),
        encoding="utf-8",
    )


def _event(
    code: str,
    entry_at: pd.Timestamp,
    *,
    gross_return: float,
    event_no: int,
) -> dict[str, object]:
    exit_at = entry_at + pd.Timedelta(days=7)
    row: dict[str, object] = {
        "signal_fact_id": f"sha256:signal-{event_no}",
        "union_signal_id": f"sha256:union-{event_no}",
        "code": code,
        "model_code": "S0000",
        "reveal_at": entry_at - pd.Timedelta(minutes=30),
        "entry_at": entry_at,
        "entry_trade_date": entry_at.date(),
        "qfq_entry_open": 10.0,
        "entry_executable": True,
        "entry_status": "OK",
        "concurrent_trigger_mask": 0x04,
        "concurrent_trigger_count": 1,
        "filter_pass_mask": 63,
        "same_code_reveal_model_count": 1,
        "same_reveal_event_count": 2,
        "amount_median_20d": 1_000_000.0 + event_no,
        "raw_entry_gap": 0.0,
        "market_regime": "UP",
        "split_id": "TRAIN",
    }
    for horizon in (5, 30, 60, 90):
        row[f"h{horizon}_status"] = "OK"
        row[f"h{horizon}_exit_at"] = exit_at
        row[f"h{horizon}_gross_return"] = gross_return
        row[f"h{horizon}_split_boundary_status"] = "AVAILABLE"
    return row


def _finalize_matrix_lineage(root: Path) -> None:
    matrix_dir = root / "matrix"
    candidates_path = matrix_dir / "development_lock_candidates.parquet"
    common = {
        "model_code": "S0000",
        "trigger_id": "ALL",
        "trigger_selector_kind": "ALL",
        "trigger_selector_value": pd.NA,
        "trigger_selector_name": "all non-zero trigger masks",
        "filter_mask": 0,
        "filter_names": "NONE",
        "filter_count": 0,
        "eligible_for_lock": True,
        "development_score": 0.60,
        "train_sample_count": 100,
        "train_net_win_rate": 0.55,
        "train_net_win_rate_ci_low": 0.50,
        "train_net_win_rate_ci_high": 0.60,
        "train_mean_net_return": 0.02,
        "train_profit_factor": 1.2,
        "train_mean_net_excess_return": 0.01,
        "validation_sample_count": 80,
        "validation_net_win_rate": 0.54,
        "validation_net_win_rate_ci_low": 0.48,
        "validation_net_win_rate_ci_high": 0.60,
        "validation_mean_net_return": 0.01,
        "validation_profit_factor": 1.1,
        "validation_mean_net_excess_return": 0.005,
    }
    pd.DataFrame(
        [
            {
                **common,
                "horizon_trading_days": horizon,
            }
            for horizon in (5, 30, 60, 90)
        ]
    ).to_parquet(candidates_path, index=False)
    features_path = root / "features" / "candidate_events.parquet"
    config_path = root / "audit" / "study_config.json"
    index_path = root / "snapshot" / "index_day.parquet"
    snapshot_manifest_path = root / "snapshot" / "manifest.json"
    development_stage_id, development_identity = _development_identity(
        features_path=features_path,
        config_path=config_path,
        index_path=index_path,
        snapshot_manifest_path=snapshot_manifest_path,
        min_train_samples=1,
        min_validation_samples=1,
        top_per_model=1,
    )
    development_manifest_path = matrix_dir / "development_manifest.json"
    write_json_atomic(
        development_manifest_path,
        {
            "study_id": "clx-30m-full-trigger-f1-f6-v1",
            "stage": "development",
            "stage_id": development_stage_id,
            "identity": development_identity,
            "data_access_contract": {
                "candidate_event_scopes_physically_read": [
                    "TRAIN",
                    "VALIDATION",
                ],
                "audit_used_in_score": False,
            },
            "outputs": {
                "lock_candidates": {
                    "path": str(candidates_path.resolve()),
                    "file_size": candidates_path.stat().st_size,
                    "file_sha256": sha256_file(candidates_path),
                }
            },
        },
    )
    run_lock(argparse.Namespace(root=root, force=True))
    lock_path = matrix_dir / "locked_config.json"
    locked = json.loads(lock_path.read_text(encoding="utf-8"))
    reveal_stage_id, reveal_identity = _reveal_identity(
        lock_path=lock_path,
        features_path=features_path,
        config_path=config_path,
        index_path=index_path,
        snapshot_manifest_path=snapshot_manifest_path,
        min_reveal_samples=1,
    )
    reveal_outputs = {
        "matrix": matrix_dir / "reveal_matrix.parquet",
        "summary": matrix_dir / "reveal_summary.csv",
        "locked_detailed": matrix_dir / "reveal_locked_detailed.parquet",
        "group_detail": matrix_dir / "reveal_locked_group_detail.parquet",
    }
    write_json_atomic(
        matrix_dir / "reveal_manifest.json",
        {
            "study_id": "clx-30m-full-trigger-f1-f6-v1",
            "stage": "reveal",
            "stage_id": reveal_stage_id,
            "lock_id": locked["lock_id"],
            "identity": reveal_identity,
            "outputs": {
                name: {
                    "path": str(path.resolve()),
                    "file_size": path.stat().st_size,
                    "file_sha256": sha256_file(path),
                }
                for name, path in reveal_outputs.items()
            },
        },
    )


def _write_fixture(root: Path, *, split_id: str = "TRAIN") -> pd.DataFrame:
    _write_lock(root)
    reveal_path = root / "matrix" / "reveal_summary.csv"
    pd.DataFrame(
        [
            {
                "selection_id": f"h{horizon}-fixture",
                "scope": "AUDIT",
                "horizon_trading_days": horizon,
                "sample_count": 40,
                "net_win_rate": 0.55,
                "net_win_rate_ci_low": 0.40,
                "net_win_rate_ci_high": 0.69,
                "mean_net_return": 0.02,
                "median_net_return": 0.01,
                "profit_factor": 1.2,
                "mean_net_excess_return": 0.01,
                "small_sample_warning": False,
            }
            for horizon in (5, 30, 60, 90)
        ]
    ).to_csv(reveal_path, index=False, encoding="utf-8-sig")
    reveal_detailed_path = root / "matrix" / "reveal_locked_detailed.parquet"
    pd.DataFrame(
        [
            {
                "scope": "AUDIT",
                "horizon_trading_days": horizon,
                "model_population": "SELECTED_MODEL",
                "aggregation": aggregation,
                "sample_count": 40,
                "net_win_rate": 0.55,
                "mean_net_return": 0.02,
                "mean_net_excess_return": 0.01,
            }
            for horizon in (5, 30, 60, 90)
            for aggregation in ("EVENT", "UNION", "MACRO", "DATE_BALANCED")
        ]
    ).to_parquet(reveal_detailed_path, index=False)
    reveal_group_path = root / "matrix" / "reveal_locked_group_detail.parquet"
    pd.DataFrame(
        [
            {
                "scope": "AUDIT",
                "horizon_trading_days": horizon,
                "model_population": "SELECTED_MODEL",
                "aggregation": "EVENT",
                "group_type": "YEAR",
                "group_value": "2024",
                "sample_count": 40,
                "net_win_rate": 0.55,
            }
            for horizon in (5, 30, 60, 90)
        ]
    ).to_parquet(reveal_group_path, index=False)
    pd.DataFrame({"fixture": [1]}).to_parquet(
        root / "matrix" / "reveal_matrix.parquet",
        index=False,
    )
    (root / "matrix" / "reveal_manifest.json").write_text(
        json.dumps(
            {
                "study_id": "clx-30m-full-trigger-f1-f6-v1",
                "stage": "reveal",
                "lock_id": "sha256:fixture-lock",
                "outputs": {
                    "summary": {
                        "path": str(reveal_path),
                        "file_size": reveal_path.stat().st_size,
                        "file_sha256": sha256_file(reveal_path),
                    },
                    "locked_detailed": {
                        "path": str(reveal_detailed_path),
                        "file_size": reveal_detailed_path.stat().st_size,
                        "file_sha256": sha256_file(reveal_detailed_path),
                    },
                    "group_detail": {
                        "path": str(reveal_group_path),
                        "file_size": reveal_group_path.stat().st_size,
                        "file_sha256": sha256_file(reveal_group_path),
                    },
                },
            }
        ),
        encoding="utf-8",
    )
    sessions = pd.bdate_range("2024-07-01", periods=12)
    index = pd.DataFrame(
        {
            "date": sessions,
            "open": 3000.0,
            "high": 3010.0,
            "low": 2990.0,
            "close": 3000.0,
        }
    )
    snapshot = root / "snapshot"
    snapshot.mkdir(parents=True)
    index_path = snapshot / "index_day.parquet"
    index.to_parquet(index_path, index=False)
    bars_dir = snapshot / "bars"
    bars_dir.mkdir()
    clocks = ("10:00", "10:30", "11:00", "11:30", "13:30", "14:00", "14:30", "15:00")
    bar_files = []
    for code, drift in (("600000", 0.02), ("600001", -0.01)):
        bars = pd.DataFrame(
            [
                {
                    "bar_at": pd.Timestamp(
                        f"{day.date().isoformat()} {clock}",
                        tz="Asia/Shanghai",
                    ),
                    "qfq_close": 10.0
                    * (
                        1
                        + drift
                        * (session_no * len(clocks) + slot)
                        / (len(sessions) * len(clocks))
                    ),
                }
                for session_no, day in enumerate(sessions)
                for slot, clock in enumerate(clocks)
            ]
        )
        bar_path = bars_dir / f"{code}.parquet"
        bars.to_parquet(bar_path, index=False)
        bar_files.append(
            {
                "code": code,
                "rows": len(bars),
                "file_size": bar_path.stat().st_size,
                "file_sha256": sha256_file(bar_path),
            }
        )
    index_source = {
        "source_code": "510980",
        "source_kind": "SHANGHAI_COMPOSITE_ETF_PROXY",
        "source_name": "上证综合ETF",
    }
    index_identity = {
        **index_source,
        "logical_path": "snapshot/index_day.parquet",
        "file_size": index_path.stat().st_size,
        "file_sha256": sha256_file(index_path),
        "rows": len(index),
    }
    audit = root / "audit"
    audit.mkdir()
    (audit / "study_config.json").write_text(
        json.dumps(
            {
                "study_id": "clx-30m-full-trigger-f1-f6-v1",
                "index_source": index_identity,
                "time_splits": {
                    "TRAIN": ["2024-01-01", "2024-07-04"],
                    "VALIDATION": ["2024-07-05", "2024-07-10"],
                    "AUDIT": ["2024-07-11", "2024-12-31"],
                },
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    (snapshot / "manifest.json").write_text(
        json.dumps(
            {
                "study_id": "clx-30m-full-trigger-f1-f6-v1",
                "snapshot_id": "sha256:fixture-snapshot",
                "index": index_identity,
                "code_files": bar_files,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    entry = pd.Timestamp("2024-07-02 10:30", tz="Asia/Shanghai")
    events = pd.DataFrame(
        [
            _event("600000", entry, gross_return=0.10, event_no=1),
            _event("600001", entry, gross_return=-0.05, event_no=2),
        ]
    )
    events["split_id"] = split_id
    events.loc[events["code"].eq("600000"), "amount_median_20d"] = 2_000_000.0
    features = root / "features"
    features.mkdir()
    event_path = features / "candidate_events.parquet"
    events.to_parquet(event_path, index=False)
    pd.DataFrame(
        [
            {
                "segment_id": "UP-0001",
                "regime": "UP",
                "start_date": sessions[0].date().isoformat(),
                "end_date": sessions[-1].date().isoformat(),
                "sessions": len(sessions),
                "start_close": 3000,
                "end_close": 3100,
                "segment_return": 0.033,
            }
        ]
    ).to_csv(features / "market_segments.csv", index=False, encoding="utf-8-sig")
    feature_summary = {
        "study_id": "clx-30m-full-trigger-f1-f6-v1",
        "candidate_event_rows": 2,
        "unique_union_signals": 2,
        "unique_stocks": 2,
        "output": {
            "path": str(event_path.resolve()),
            "file_size": event_path.stat().st_size,
            "file_sha256": sha256_file(event_path),
        },
    }
    (features / "summary.json").write_text(
        json.dumps(feature_summary),
        encoding="utf-8",
    )
    signal_set_id = "sha256:" + "1" * 64
    (features / "manifest.json").write_text(
        json.dumps(
            {
                "study_id": "clx-30m-full-trigger-f1-f6-v1",
                "snapshot_id": "sha256:fixture-snapshot",
                "signal_set_id": signal_set_id,
                "summary": feature_summary,
            }
        ),
        encoding="utf-8",
    )
    replay = root / "replay"
    replay.mkdir()
    (replay / "manifest.json").write_text(
        json.dumps(
            {
                "study_id": "clx-30m-full-trigger-f1-f6-v1",
                "snapshot_id": "sha256:fixture-snapshot",
                "signal_set_id": signal_set_id,
            }
        ),
        encoding="utf-8",
    )
    _finalize_matrix_lineage(root)
    return events


def test_lock_rejects_filter_mask_outside_six_bits(tmp_path: Path) -> None:
    with pytest.raises(PortfolioContractError, match="0..63"):
        _parse_locked_selections(_locked_payload(filter_mask=64))


def test_lock_rejects_another_study_identity(tmp_path: Path) -> None:
    payload = _locked_payload()
    payload["study_id"] = "another-study"
    path = tmp_path / "matrix" / "locked_config.json"
    path.parent.mkdir(parents=True)
    path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(PortfolioContractError, match="study_id mismatch"):
        load_locked_selections(path)


def test_load_lock_rejects_tampered_selection_with_old_lock_id(
    tmp_path: Path,
) -> None:
    _write_fixture(tmp_path)
    path = tmp_path / "matrix" / "locked_config.json"
    payload = json.loads(path.read_text(encoding="utf-8"))
    payload["selections"][0]["model_code"] = "S0001"
    path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(PortfolioContractError, match="lock_id mismatch"):
        load_locked_selections(path)


def test_selection_applies_single_trigger_and_f1_f6_mask(tmp_path: Path) -> None:
    selection = _parse_locked_selections(_locked_payload())[0]
    selection = selection.__class__(
        **{
            **selection.__dict__,
            "trigger_kind": "SINGLE_BIT",
            "trigger_value": 0x04,
            "filter_mask": 0b100001,
            "filter_names": (
                "F1",
                "F6",
            ),
        }
    )
    entry = pd.Timestamp("2024-07-02 10:30", tz="Asia/Shanghai")
    passing = _event("600000", entry, gross_return=0.1, event_no=1)
    failing = _event("600001", entry, gross_return=0.1, event_no=2)
    failing["concurrent_trigger_mask"] = 0x02
    frame = pd.DataFrame([passing, failing])

    actual = select_locked_candidates(frame, selection)

    assert actual["code"].tolist() == ["600000"]
    assert math.isclose(actual.loc[0, "qfq_exit_open"], 11.0)


@pytest.mark.parametrize("horizon", (5, 30, 60, 90))
def test_selection_scope_contract_is_fail_closed_for_every_horizon(
    tmp_path: Path,
    horizon: int,
) -> None:
    selection = next(
        item
        for item in _parse_locked_selections(_locked_payload())
        if item.horizon == horizon
    )
    entry = pd.Timestamp("2024-07-02 10:30", tz="Asia/Shanghai")
    train_available = _event("600000", entry, gross_return=0.1, event_no=1)
    validation_available = _event("600001", entry, gross_return=0.1, event_no=2)
    validation_available["split_id"] = "VALIDATION"
    audit_available = _event("600002", entry, gross_return=0.1, event_no=3)
    audit_available["split_id"] = "AUDIT"
    train_purged = _event("600003", entry, gross_return=0.1, event_no=4)
    train_purged[f"h{horizon}_split_boundary_status"] = "PURGED"
    audit_purged = _event("600004", entry, gross_return=0.1, event_no=5)
    audit_purged["split_id"] = "AUDIT"
    audit_purged[f"h{horizon}_split_boundary_status"] = "PURGED"
    missing_boundary = _event("600005", entry, gross_return=0.1, event_no=6)
    missing_boundary[f"h{horizon}_split_boundary_status"] = pd.NA
    unknown_boundary = _event("600006", entry, gross_return=0.1, event_no=7)
    unknown_boundary[f"h{horizon}_split_boundary_status"] = "UNKNOWN"
    unknown_split = _event("600007", entry, gross_return=0.1, event_no=8)
    unknown_split["split_id"] = "UNKNOWN"
    missing_split = _event("600008", entry, gross_return=0.1, event_no=9)
    missing_split["split_id"] = pd.NA
    events = pd.DataFrame(
        [
            train_available,
            validation_available,
            audit_available,
            train_purged,
            audit_purged,
            missing_boundary,
            unknown_boundary,
            unknown_split,
            missing_split,
        ]
    )

    available = select_locked_candidates(events, selection, scope="AVAILABLE")
    audit = select_locked_candidates(events, selection, scope="AUDIT")

    assert set(available["code"]) == {"600000", "600001", "600002"}
    assert audit["code"].tolist() == ["600002"]

    with pytest.raises(PortfolioContractError, match="split_id"):
        select_locked_candidates(
            events.drop(columns="split_id"),
            selection,
            scope="AVAILABLE",
        )
    with pytest.raises(
        PortfolioContractError,
        match=f"h{horizon}_split_boundary_status",
    ):
        select_locked_candidates(
            events.drop(columns=f"h{horizon}_split_boundary_status"),
            selection,
            scope="AVAILABLE",
        )


def test_checkpoint_keys_isolate_available_audit_and_matched90(
    tmp_path: Path,
) -> None:
    selection = _parse_locked_selections(_locked_payload())[0]

    keys = {
        scope: _checkpoint_key(
            selection=selection,
            scope=scope,
            daily_entry_limit=5,
            ranking_policy="quality",
            random_seed=None,
        )
        for scope in ("AVAILABLE", "AUDIT", "MATCHED90")
    }

    assert len(set(keys.values())) == 3
    assert "_available_" in keys["AVAILABLE"]
    assert "_audit_" in keys["AUDIT"]
    assert "_matched90_" in keys["MATCHED90"]
    with pytest.raises(PortfolioContractError, match="unsupported checkpoint scope"):
        _checkpoint_key(
            selection=selection,
            scope="UNKNOWN",
            daily_entry_limit=5,
            ranking_policy="quality",
            random_seed=None,
        )


def test_run_rejects_reveal_artifact_changed_after_manifest(
    tmp_path: Path,
) -> None:
    _write_fixture(tmp_path)
    path = tmp_path / "matrix" / "reveal_summary.csv"
    path.write_text(path.read_text(encoding="utf-8-sig") + "\n", encoding="utf-8-sig")

    with pytest.raises(PortfolioContractError, match="summary identity mismatch"):
        run_portfolios(
            root=tmp_path,
            random_seeds=1,
            include_audit_scope=False,
        )


def test_run_rejects_old_reveal_manifest_mixed_with_current_lock(
    tmp_path: Path,
) -> None:
    _write_fixture(tmp_path)
    path = tmp_path / "matrix" / "reveal_manifest.json"
    payload = json.loads(path.read_text(encoding="utf-8"))
    payload["lock_id"] = "sha256:" + "0" * 64
    path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(PortfolioContractError, match="reveal identity mismatch"):
        run_portfolios(
            root=tmp_path,
            random_seeds=0,
            include_audit_scope=False,
        )


def test_snapshot_contract_verifies_every_declared_file_and_identity(
    tmp_path: Path,
) -> None:
    _write_fixture(tmp_path)
    benchmark = {
        "source_code": "510980",
        "source_kind": "SHANGHAI_COMPOSITE_ETF_PROXY",
        "source_name": "上证综合ETF",
    }

    verified = validate_snapshot_inputs(
        tmp_path,
        index_benchmark=benchmark,
    )

    assert verified["status"] == "VERIFIED"
    assert verified["all_declared_files_verified"] is True
    assert verified["verified_bar_file_count"] == 2
    assert verified["verified_file_count"] == 3

    for relative_path in (
        "snapshot/index_day.parquet",
        "snapshot/bars/600000.parquet",
    ):
        path = tmp_path / relative_path
        original = path.read_bytes()
        path.write_bytes(original + b"x")
        with pytest.raises(PortfolioContractError, match="size mismatch"):
            validate_snapshot_inputs(tmp_path, index_benchmark=benchmark)
        path.write_bytes(original)


def test_snapshot_contract_rejects_declared_bar_rows_drift(tmp_path: Path) -> None:
    _write_fixture(tmp_path)
    path = tmp_path / "snapshot" / "manifest.json"
    payload = json.loads(path.read_text(encoding="utf-8"))
    payload["code_files"][0]["rows"] += 1
    path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(PortfolioContractError, match="Parquet row count mismatch"):
        validate_snapshot_inputs(
            tmp_path,
            index_benchmark={
                "source_code": "510980",
                "source_kind": "SHANGHAI_COMPOSITE_ETF_PROXY",
                "source_name": "上证综合ETF",
            },
        )


@pytest.mark.parametrize(
    ("mutation", "message"),
    (
        ("study", "study_id mismatch"),
        ("index_path", "index logical_path"),
        ("bar_path", "path mismatch"),
        ("duplicate_code", "duplicate code"),
    ),
)
def test_snapshot_contract_rejects_ambiguous_manifest_identity(
    tmp_path: Path,
    mutation: str,
    message: str,
) -> None:
    _write_fixture(tmp_path)
    path = tmp_path / "snapshot" / "manifest.json"
    payload = json.loads(path.read_text(encoding="utf-8"))
    if mutation == "study":
        payload["study_id"] = "another-study"
    elif mutation == "index_path":
        payload["index"]["logical_path"] = "../index_day.parquet"
    elif mutation == "bar_path":
        payload["code_files"][0]["path"] = "../600000.parquet"
    else:
        payload["code_files"].append(dict(payload["code_files"][0]))
    path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(PortfolioContractError, match=message):
        validate_snapshot_inputs(
            tmp_path,
            index_benchmark={
                "source_code": "510980",
                "source_kind": "SHANGHAI_COMPOSITE_ETF_PROXY",
                "source_name": "上证综合ETF",
            },
        )


def test_run_requires_feature_summary_identity(tmp_path: Path) -> None:
    _write_fixture(tmp_path)
    (tmp_path / "features" / "summary.json").unlink()

    with pytest.raises(
        PortfolioContractError,
        match="summary.json",
    ):
        run_portfolios(
            root=tmp_path,
            random_seeds=0,
            include_audit_scope=False,
        )


def test_run_rejects_features_manifest_signal_set_drift(tmp_path: Path) -> None:
    _write_fixture(tmp_path)
    path = tmp_path / "features" / "manifest.json"
    payload = json.loads(path.read_text(encoding="utf-8"))
    payload["signal_set_id"] = "sha256:" + "2" * 64
    path.write_text(json.dumps(payload), encoding="utf-8")

    with pytest.raises(PortfolioContractError, match="signal_set_id lineage mismatch"):
        run_portfolios(
            root=tmp_path,
            random_seeds=0,
            include_audit_scope=False,
        )


@pytest.mark.parametrize(
    "relative_path",
    ("features/market_segments.csv",),
)
def test_run_id_changes_when_bound_input_changes(
    tmp_path: Path,
    relative_path: str,
) -> None:
    _write_fixture(tmp_path)
    first = run_portfolios(
        root=tmp_path,
        random_seeds=0,
        include_audit_scope=False,
    )
    path = tmp_path / relative_path
    path.write_bytes(path.read_bytes() + b"\n")

    second = run_portfolios(
        root=tmp_path,
        random_seeds=0,
        include_audit_scope=False,
    )

    assert second["reused"] is False
    assert second["run_id"] != first["run_id"]


def test_run_rejects_snapshot_manifest_drift_from_matrix_lock(tmp_path: Path) -> None:
    _write_fixture(tmp_path)
    path = tmp_path / "snapshot" / "manifest.json"
    path.write_bytes(path.read_bytes() + b"\n")

    with pytest.raises(PortfolioContractError, match="matrix lock lineage mismatch"):
        run_portfolios(
            root=tmp_path,
            random_seeds=0,
            include_audit_scope=False,
        )


def test_rebuild_report_rejects_external_input_drift(tmp_path: Path) -> None:
    _write_fixture(tmp_path)
    run_portfolios(
        root=tmp_path,
        random_seeds=0,
        include_audit_scope=False,
    )
    path = tmp_path / "features" / "market_segments.csv"
    path.write_bytes(path.read_bytes() + b"\n")

    with pytest.raises(
        PortfolioContractError,
        match="frozen run_contract external input drift: market_segments",
    ):
        rebuild_report(tmp_path)


def test_simulator_limits_daily_entries_and_charges_both_sides(
    tmp_path: Path,
) -> None:
    events = _write_fixture(tmp_path)
    selection = load_locked_selections(tmp_path / "matrix" / "locked_config.json")[0]
    candidates = select_locked_candidates(events, selection)
    marks = MarkStore(tmp_path)
    marks.load(candidates["code"])
    clock = build_simulation_clock(tmp_path, [candidates], marks)

    actual = simulate_portfolio(
        candidates,
        selection=selection,
        scope="AVAILABLE",
        clock=clock,
        marks=marks,
        daily_entry_limit=1,
        ranking_policy="quality",
    )

    assert actual.summary["closed_trades"] == 1
    assert actual.summary["rejected_daily_limit"] == 1
    assert actual.summary["final_equity"] > INITIAL_CAPITAL
    trade = actual.trades.iloc[0]
    expected = (1 + 0.10) * (1 - FEE_PER_SIDE) / (1 + FEE_PER_SIDE) - 1
    assert math.isclose(trade["net_return"], expected)
    assert actual.summary["total_fees"] > 0


def test_smoke_run_writes_machine_report_and_excel(tmp_path: Path) -> None:
    _write_fixture(tmp_path)

    result = run_portfolios(
        root=tmp_path,
        random_seeds=3,
        include_audit_scope=False,
    )

    output = Path(result["portfolio_dir"])
    summary = pd.read_parquet(output / "portfolio_summary.parquet")
    random_runs = pd.read_parquet(output / "random_order_runs.parquet")
    report = (output / "report.md").read_text(encoding="utf-8")
    manifest = json.loads((output / "manifest.json").read_text(encoding="utf-8"))
    config = json.loads((output / "portfolio_config.json").read_text(encoding="utf-8"))
    assert len(summary) == 4 * 6
    assert len(random_runs) == 4 * 6 * 3
    assert set(summary["daily_entry_limit"]) == {
        "1",
        "3",
        "5",
        "10",
        "20",
        "UNLIMITED",
    }
    assert "F1-F6 共64个子集" in report
    assert "样本外 AUDIT 一次性揭示" in report
    assert "AUDIT多聚合口径" in report
    assert "|5|40|55.00%|" in report
    assert "滑点、印花税、最低佣金、100股取整" in report
    workbook_path = output / "clx_30m_portfolio_report.xlsx"
    assert workbook_path.is_file()
    workbook = load_workbook(workbook_path)
    assert "Charts" in workbook.sheetnames
    assert "RevealDetailed" in workbook.sheetnames
    assert "RevealByPeriod" in workbook.sheetnames
    assert len(workbook["Charts"]._charts) == 2
    portfolio_sheet = workbook["Portfolio"]
    headers = {cell.value: cell.column for cell in portfolio_sheet[1]}
    assert portfolio_sheet.freeze_panes == "A2"
    assert portfolio_sheet.auto_filter.ref == portfolio_sheet.dimensions
    assert portfolio_sheet["A1"].fill.fgColor.rgb.endswith("1F4E78")
    assert portfolio_sheet["A1"].font.bold is True
    assert portfolio_sheet["A1"].font.color.rgb.endswith("FFFFFF")
    assert portfolio_sheet.column_dimensions["A"].width == 24
    assert all(
        dimension.width <= 42
        for dimension in portfolio_sheet.column_dimensions.values()
        if dimension.width is not None
    )
    assert portfolio_sheet.cell(2, headers["total_return"]).number_format == "0.00%"
    assert (
        portfolio_sheet.cell(2, headers["initial_capital"]).number_format
        == '"¥"#,##0.00'
    )
    assert (
        portfolio_sheet.cell(2, headers["candidate_signals"]).number_format == "#,##0"
    )
    assert portfolio_sheet.cell(2, headers["average_positions"]).number_format == "0.00"
    assert portfolio_sheet.cell(2, headers["average_win"]).number_format == "0.00%"
    assert (
        portfolio_sheet.cell(2, headers["start_at"]).number_format == "yyyy-mm-dd hh:mm"
    )
    for chart in workbook["Charts"]._charts:
        assert chart.y_axis.numFmt.formatCode == "0.00%"
        assert [series.tx.v for series in chart.series] == [
            "5日",
            "30日",
            "60日",
            "90日",
        ]
    chart_data_sheet = workbook["ChartData"]
    chart_data_headers = {cell.value: cell.column for cell in chart_data_sheet[1]}
    assert chart_data_sheet["B2"].number_format == "0.00%"
    for horizon in (5, 30, 60, 90):
        column = chart_data_headers[f"h{horizon}_normalized_equity"]
        assert chart_data_sheet.cell(2, column).number_format == "0.00%"
    date_label_column = chart_data_headers["trade_date_label"]
    trade_date_column = chart_data_headers["trade_date"]
    assert chart_data_sheet.cell(2, date_label_column).value == pd.Timestamp(
        chart_data_sheet.cell(2, trade_date_column).value
    ).strftime("%Y-%m-%d")
    date_label_letter = get_column_letter(date_label_column)
    expected_categories = (
        f"'ChartData'!${date_label_letter}$2:"
        f"${date_label_letter}${chart_data_sheet.max_row}"
    )
    for chart in workbook["Charts"]._charts:
        for series in chart.series:
            assert series.cat.strRef is not None
            assert series.cat.strRef.f == expected_categories
            assert series.cat.numRef is None
    assert workbook["Charts"].sheet_view.showGridLines is False
    contract_sheet = workbook["Contract"]
    assert contract_sheet.column_dimensions["B"].width <= 42
    for row in range(2, 8):
        contract_cell = contract_sheet.cell(row, 2)
        assert contract_cell.alignment.wrap_text is True
        assert contract_cell.alignment.vertical == "top"
        assert contract_sheet.row_dimensions[row].height >= 30
    assert "冻结配置的资金 AUDIT" in report
    assert "3组敏感性" in report
    assert "3组SHA随机排序分位数" in report
    assert "100组" not in report
    assert manifest["selection_count"] == 4
    assert manifest["random_portfolios"] == 72
    assert "510980上证综合ETF代理" in report
    assert "相对510980上证综合ETF代理平均超额" in report
    assert "相对上证平均超额" not in report
    assert config["index_benchmark"] == manifest["index_benchmark"]
    assert config["index_benchmark"]["benchmark_label"] == "510980上证综合ETF代理"
    assert config["filter_descriptions"]["F6"].startswith("510980上证综合ETF代理")
    assert config["snapshot_verification"] == manifest["snapshot_verification"]
    assert config["snapshot_verification"]["all_declared_files_verified"] is True
    assert {
        item["logical_path"] for item in manifest["outputs"]
    } == REQUIRED_LOGICAL_OUTPUTS
    reproduce_command = (output / "reproduce_command.txt").read_text(encoding="utf-8")
    assert "--no-audit-scope" in reproduce_command
    assert f'"{REPRODUCE_SCRIPT}"' in reproduce_command
    assert str(Path(__file__).resolve().parents[3]) not in reproduce_command
    assert (
        manifest["portfolio_logic_sha256"]
        == config["run_contract"]["portfolio_logic_sha256"]
        == sha256_file(
            Path(__file__).resolve().parents[3]
            / "script"
            / "clx_backtest"
            / "research"
            / "clx_30m_f1_f6_portfolio.py"
        )
    )
    assert str(manifest["portfolio_logic_sha256"]) in report
    for name in (
        "snapshot_manifest",
        "market_segments",
        "feature_summary",
        "study_config",
    ):
        assert set(config["run_contract"][name]) == {
            "logical_path",
            "file_size",
            "sha256",
        }
        assert manifest["input"][name] == config["run_contract"][name]
    bar_path = tmp_path / "snapshot" / "bars" / "600000.parquet"
    original_bar = bar_path.read_bytes()
    bar_path.write_bytes(original_bar + b"x")
    with pytest.raises(PortfolioContractError, match="size mismatch"):
        run_portfolios(
            root=tmp_path,
            random_seeds=3,
            include_audit_scope=False,
        )
    bar_path.write_bytes(original_bar)
    reused = run_portfolios(
        root=tmp_path,
        random_seeds=3,
        include_audit_scope=False,
    )
    assert reused["reused"] is True
    assert reused["run_id"] == manifest["run_id"]

    manifest_path = output / "manifest.json"
    original_manifest = manifest_path.read_bytes()
    original_outputs = manifest["outputs"]
    invalid_outputs: tuple[list[dict[str, object]], ...] = (
        [],
        original_outputs[:-1],
        [*original_outputs, dict(original_outputs[0])],
    )
    for outputs in invalid_outputs:
        invalid_manifest = {**manifest, "outputs": outputs}
        manifest_path.write_text(json.dumps(invalid_manifest), encoding="utf-8")
        with pytest.raises(PortfolioContractError, match="outputs"):
            _completed_result(tmp_path, run_id=manifest["run_id"])
    manifest_path.write_bytes(original_manifest)
    invalid_manifest = {**manifest, "portfolio_logic_sha256": "0" * 64}
    manifest_path.write_text(json.dumps(invalid_manifest), encoding="utf-8")
    with pytest.raises(PortfolioContractError, match="logic SHA256 mismatch"):
        _completed_result(tmp_path, run_id=manifest["run_id"])
    manifest_path.write_bytes(original_manifest)

    config_path = output / "portfolio_config.json"
    original_config = config_path.read_bytes()
    invalid_config = json.loads(original_config.decode("utf-8"))
    invalid_config["run_contract"]["portfolio_logic_sha256"] = "0" * 64
    config_path.write_text(json.dumps(invalid_config), encoding="utf-8")
    with pytest.raises(PortfolioContractError, match="logic SHA256 mismatch"):
        rebuild_report(tmp_path)
    config_path.write_bytes(original_config)
    assert Path(rebuild_report(tmp_path)["report"]).is_file()

    checkpoint = next((output / "ckpt").rglob("complete.json"))
    checkpoint_mtime = checkpoint.stat().st_mtime_ns
    (output / "manifest.json").unlink()
    resumed = run_portfolios(
        root=tmp_path,
        random_seeds=3,
        include_audit_scope=False,
    )
    assert resumed["reused"] is False
    assert checkpoint.stat().st_mtime_ns == checkpoint_mtime


def test_audit_report_uses_current_metrics_without_frozen_conclusions(
    tmp_path: Path,
) -> None:
    _write_fixture(tmp_path, split_id="AUDIT")

    result = run_portfolios(
        root=tmp_path,
        random_seeds=0,
        include_audit_scope=True,
    )

    output = Path(result["portfolio_dir"])
    report = (output / "report.md").read_text(encoding="utf-8")
    summary = pd.read_parquet(output / "portfolio_summary.parquet")
    audit_five = summary[
        summary["scope"].eq("AUDIT")
        & summary["horizon_trading_days"].eq(5)
        & summary["daily_entry_limit"].eq("5")
        & summary["ranking_policy"].eq("quality")
    ].iloc[0]
    assert f"{audit_five.total_return:.2%}" in report
    assert "35%" not in report
    assert "2026Q2" not in report
    assert "收益集中于震荡段" not in report
    assert "仅覆盖2026年" not in report
    assert "2024年以来短样本" not in report


def test_audit_interpretation_switches_with_current_metrics() -> None:
    positive = pd.Series(
        {
            "total_return": 0.12,
            "profit_factor": 1.4,
            "closed_win_rate": 0.58,
        }
    )
    negative = pd.Series(
        {
            "total_return": -0.08,
            "profit_factor": 0.7,
            "closed_win_rate": 0.42,
        }
    )
    negative_high_pf = pd.Series(
        {
            "total_return": -0.04,
            "profit_factor": 1.3,
            "closed_win_rate": 0.47,
        }
    )

    assert "总收益与PF均高于各自基准" in _portfolio_quality_statement(positive)
    assert "总收益与PF均未高于各自基准" in _portfolio_quality_statement(negative)
    assert "资金方向一致" in _stability_statement(positive, positive)
    assert "资金方向不一致" in _stability_statement(positive, negative)
    both_negative = _stability_statement(negative, negative_high_pf)
    assert "两段总收益均为负" in both_negative
    assert "AUDIT PF高于1" in both_negative
    assert "AUDIT PF未高于1" not in both_negative


def test_excel_temporary_file_is_removed_when_replace_fails(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    def fail_replace(_source: Path, _destination: Path) -> None:
        raise OSError("fixture replace failure")

    monkeypatch.setattr(
        "script.clx_backtest.research.clx_30m_f1_f6_portfolio.os.replace",
        fail_replace,
    )
    path = tmp_path / "report.xlsx"
    empty = pd.DataFrame()

    with pytest.raises(OSError, match="fixture replace failure"):
        _write_excel(
            path,
            summary=empty,
            random_summary=empty,
            period_metrics=empty,
            daily_curves=empty,
            chart_data=empty,
            locked=empty,
            baseline=empty,
            reveal=empty,
            reveal_detailed=empty,
            reveal_groups=empty,
        )

    assert not path.with_name(".report.tmp.xlsx").exists()
